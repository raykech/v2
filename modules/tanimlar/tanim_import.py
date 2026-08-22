# -*- coding: utf-8 -*-
"""
Tanımlar (Cari, Stok, Hizmet Kartı) Excel import yardımcıları.

- Örnek Excel şablonu oluşturur
- Excel dosyasını okur
- Satırları doğrular ve kaydedilmeye hazır kart listesine çevirir
- core.services.kaydet_kart ile kaydedilmeye hazır veri döndürür
"""
from datetime import datetime, date
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    load_workbook = None
    Alignment = Font = PatternFill = DataValidation = None

from core.db import veritabani_baglan
from core.services import kaydet_kart

CARi_KOLONLAR = ["Unvan", "Tür", "Telefon", "Durum"]
STOK_KOLONLAR = ["Stok Adı", "Stok Kodu", "Kategori", "Birim", "Alış Fiyatı", "Satış Fiyatı", "KDV %", "Kritik Miktar", "Durum"]
HIZMET_KOLONLAR = ["Kart Adı", "Tür", "Grup", "KDV %", "Durum"]
KASA_KOLONLAR = ["Kasa Adı", "Durum"]
BANKA_KURUM_KOLONLAR = ["Kurum Adı", "Durum"]
BANKA_HESAP_KOLONLAR = ["Hesap Adı", "Banka Kurumu", "Hesap Türü", "IBAN", "Komisyon %", "Durum"]

SHEET_KOLONLAR = {
    "Cari Kartlar": CARi_KOLONLAR,
    "Stok Kartları": STOK_KOLONLAR,
    "Hizmet Kartları": HIZMET_KOLONLAR,
    "Kasa Kartları": KASA_KOLONLAR,
    "Banka Kurumları": BANKA_KURUM_KOLONLAR,
    "Banka Hesapları": BANKA_HESAP_KOLONLAR,
}


def _metin(deger):
    if deger is None:
        return ""
    if isinstance(deger, float) and pd is not None and pd.isna(deger):
        return ""
    if isinstance(deger, str):
        return deger.strip()
    if isinstance(deger, (datetime, date)):
        return deger.strftime("%d.%m.%Y")
    return str(deger).strip()


def _sayi(deger):
    if deger is None:
        return None
    if isinstance(deger, bool):
        return None
    if isinstance(deger, (int, float)):
        if pd is not None and isinstance(deger, float) and pd.isna(deger):
            return None
        return float(deger)

    s = str(deger).strip().replace(" ", "").replace("TL", "").replace("₺", "").replace("%", "")
    if not s:
        return None

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    if s.count(".") > 1:
        s = s.replace(".", "")
    elif s.count(".") == 1 and len(s.split(".")[1]) == 3:
        s = s.replace(".", "")

    try:
        return float(s)
    except ValueError:
        return None


def _durum_to_int(durum_str):
    durum_str = _metin(durum_str)
    if durum_str in ("Pasif", "0", "False"):
        return 0
    return 1


def tanim_ornek_excel_olustur(dosya_yolu):
    """Cari, Stok ve Hizmet kartları için örnek Excel şablonu oluşturur."""
    if pd is None:
        raise RuntimeError("Excel şablonu oluşturmak için 'pandas' ve 'openpyxl' gereklidir.")

    cari_data = [
        ["Örnek Müşteri", "Müşteri", "0555 000 00 00", "Aktif"],
        ["Örnek Tedarikçi", "Tedarikçi", "", "Aktif"],
    ]
    stok_data = [
        ["Örnek Ürün", "UR-001", "Örnek Kategori", "Adet", 100, 150, 20, 10, "Aktif"],
        ["Örnek Hammadde", "HM-001", "", "kg", 50, 0, 20, 0, "Aktif"],
    ]
    hizmet_data = [
        ["Kira Gideri", "Gider", "Diğer", 20, "Aktif"],
        ["Satış Geliri", "Gelir", "Diğer", 20, "Aktif"],
    ]
    kasa_data = [
        ["Ana Kasa", "Aktif"],
        ["Yedek Kasa", "Aktif"],
    ]
    banka_kurum_data = [
        ["Örnek Banka A", "Aktif"],
        ["Örnek Banka B", "Aktif"],
    ]
    banka_hesap_data = [
        ["Vadesiz Hesap", "Örnek Banka A", "Vadesiz", "TR12 3456 7890 1234 5678 9000", 0, "Aktif"],
        ["POS Terminal", "Örnek Banka B", "POS", "", 1.5, "Aktif"],
    ]

    with pd.ExcelWriter(dosya_yolu, engine="openpyxl") as writer:
        pd.DataFrame(cari_data, columns=CARi_KOLONLAR).to_excel(writer, sheet_name="Cari Kartlar", index=False)
        pd.DataFrame(stok_data, columns=STOK_KOLONLAR).to_excel(writer, sheet_name="Stok Kartları", index=False)
        pd.DataFrame(hizmet_data, columns=HIZMET_KOLONLAR).to_excel(writer, sheet_name="Hizmet Kartları", index=False)
        pd.DataFrame(kasa_data, columns=KASA_KOLONLAR).to_excel(writer, sheet_name="Kasa Kartları", index=False)
        pd.DataFrame(banka_kurum_data, columns=BANKA_KURUM_KOLONLAR).to_excel(writer, sheet_name="Banka Kurumları", index=False)
        pd.DataFrame(banka_hesap_data, columns=BANKA_HESAP_KOLONLAR).to_excel(writer, sheet_name="Banka Hesapları", index=False)

    if load_workbook is None:
        return

    wb = load_workbook(dosya_yolu)
    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_fill = PatternFill(start_color="4A69BD", end_color="4A69BD", fill_type="solid")

    for ws in wb.worksheets:
        for hucre in ws[1]:
            hucre.font = baslik_font
            hucre.fill = baslik_fill
            hucre.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 12

    aciklama_ws = wb.create_sheet("Açıklama", 0)
    aciklama_ws.column_dimensions["A"].width = 120
    satirlar = [
        ("TANIMLAR EXCEL İMPORT REHBERİ", True),
        ("", False),
        ("Bu dosyada 3 sayfa bulunur: Cari Kartlar, Stok Kartları, Hizmet Kartları.", False),
        ("Her sayfadaki örnek satırları kendi verilerinizle değiştirin.", False),
        ("Aynı isimde kart zaten tanımlıysa import hata verir; önce mevcut kartı düzenleyin.", False),
        ("Stok Kodu boş bırakılırsa otomatik oluşturulur.", False),
        ("Hizmet kartında Grup boş bırakılırsa 'Diğer' grubu kullanılır.", False),
        ("KDV % boş bırakılırsa hizmet kartlarında 20, stok kartlarında 20 varsayılır.", False),
        ("", False),
        ("CARİ SAYFASI:", True),
        ("- Unvan zorunludur.", False),
        ("- Tür: Müşteri / Tedarikçi / Diğer (boşsa Müşteri)", False),
        ("- Telefon opsiyoneldir.", False),
        ("", False),
        ("STOK SAYFASI:", True),
        ("- Stok Adı zorunludur.", False),
        ("- Kategori ve Birim boşsa kart boş olarak kaydedilir.", False),
        ("- Alış/Satış fiyatı, KDV %, Kritik Miktar opsiyoneldir.", False),
        ("", False),
        ("HİZMET SAYFASI:", True),
        ("- Kart Adı zorunludur.", False),
        ("- Tür: Gider / Gelir (boşsa Gider)", False),
        ("- Grup boşsa 'Diğer' grubu kullanılır.", False),
    ]
    for i, (metin, kalin) in enumerate(satirlar, start=1):
        hucre = aciklama_ws.cell(row=i, column=1, value=metin)
        if kalin:
            hucre.font = Font(bold=True, size=12)

    wb.save(dosya_yolu)


def tanim_excel_oku(dosya_yolu):
    """
    Excel'deki tanım sayfalarını okur.

    Dönüş: {"Cari Kartlar": [...], "Stok Kartları": [...], "Hizmet Kartları": [...]}
    """
    if pd is None:
        raise RuntimeError("Excel okumak için 'pandas' ve 'openpyxl' gereklidir.")

    sayfalar = pd.read_excel(dosya_yolu, sheet_name=None)
    sonuc = {}

    for sayfa_adi, df in sayfalar.items():
        # Bilinen şema hangisiyle eşleşiyor?
        secili_kolonlar = None
        secili_tip = None
        for tip, kolonlar in SHEET_KOLONLAR.items():
            if all(kolon in df.columns for kolon in kolonlar):
                secili_kolonlar = kolonlar
                secili_tip = tip
                break

        if secili_kolonlar is None:
            continue

        df = df.dropna(how="all")
        satirlar = []
        for _, row in df.iterrows():
            satir = {}
            for kolon in secili_kolonlar:
                satir[kolon] = row.get(kolon)
            satirlar.append(satir)
        sonuc[secili_tip] = satirlar

    return sonuc


def _cari_duplicate_kontrol(cursor, unvan, firma_id, mevcut_isimler):
    if unvan in mevcut_isimler:
        return f"'{unvan}' bu dosyada birden fazla kez girilmiş."
    cursor.execute("SELECT id FROM cariler WHERE firma_id=? AND unvan=?", (firma_id, unvan))
    if cursor.fetchone():
        return f"'{unvan}' adında bir cari zaten tanımlı."
    return None


def _stok_duplicate_kontrol(cursor, stok_adi, stok_kodu, firma_id, mevcut_adlar, mevcut_kodlar):
    if stok_adi in mevcut_adlar:
        return f"'{stok_adi}' bu dosyada birden fazla kez girilmiş."
    cursor.execute("SELECT id FROM stoklar WHERE firma_id=? AND stok_adi=?", (firma_id, stok_adi))
    if cursor.fetchone():
        return f"'{stok_adi}' adında bir stok zaten tanımlı."

    if stok_kodu:
        if stok_kodu in mevcut_kodlar:
            return f"'{stok_kodu}' stok kodu bu dosyada birden fazla kez girilmiş."
        cursor.execute("SELECT id FROM stoklar WHERE firma_id=? AND stok_kodu=?", (firma_id, stok_kodu))
        if cursor.fetchone():
            return f"'{stok_kodu}' stok kodu zaten tanımlı."
    return None


def _hizmet_duplicate_kontrol(cursor, kart_adi, firma_id, mevcut_adlar):
    if kart_adi in mevcut_adlar:
        return f"'{kart_adi}' bu dosyada birden fazla kez girilmiş."
    cursor.execute("SELECT id FROM hizmet_kartlari WHERE firma_id=? AND kart_adi=?", (firma_id, kart_adi))
    if cursor.fetchone():
        return f"'{kart_adi}' adında bir hizmet kartı zaten tanımlı."
    return None


def _kasa_duplicate_kontrol(cursor, kasa_adi, firma_id, mevcut_adlar):
    if kasa_adi in mevcut_adlar:
        return f"'{kasa_adi}' bu dosyada birden fazla kez girilmiş."
    cursor.execute("SELECT id FROM kasalar WHERE firma_id=? AND kasa_adi=?", (firma_id, kasa_adi))
    if cursor.fetchone():
        return f"'{kasa_adi}' adında bir kasa zaten tanımlı."
    return None


def _banka_kurum_duplicate_kontrol(cursor, kurum_adi, firma_id, mevcut_adlar):
    if kurum_adi in mevcut_adlar:
        return f"'{kurum_adi}' bu dosyada birden fazla kez girilmiş."
    cursor.execute("SELECT id FROM banka_kurumlari WHERE firma_id=? AND kurum_adi=?", (firma_id, kurum_adi))
    if cursor.fetchone():
        return f"'{kurum_adi}' adında bir banka kurumu zaten tanımlı."
    return None


def _banka_hesap_duplicate_kontrol(cursor, hesap_adi, firma_id, mevcut_adlar):
    if hesap_adi in mevcut_adlar:
        return f"'{hesap_adi}' bu dosyada birden fazla kez girilmiş."
    cursor.execute("SELECT id FROM banka_hesaplari WHERE firma_id=? AND hesap_adi=?", (firma_id, hesap_adi))
    if cursor.fetchone():
        return f"'{hesap_adi}' adında bir banka hesabı zaten tanımlı."
    return None


def _kurum_id_bul_veya_olustur(cursor, kurum_adi, firma_id):
    """Banka kurumunu bulur; yoksa oluşturur."""
    kurum_adi = kurum_adi or "Diğer"
    cursor.execute("SELECT id FROM banka_kurumlari WHERE firma_id=? AND kurum_adi=?", (firma_id, kurum_adi))
    row = cursor.fetchone()
    if row:
        return row[0]
    return kaydet_kart(
        cursor,
        "banka_kurumlari",
        {"kurum_adi": kurum_adi, "firma_id": firma_id, "durum": 1},
    )


def _kategori_birim_id(cursor, grup, deger, firma_id):
    """genel_tanimlar içinde yoksa oluşturur ve ID döndürür."""
    if not deger:
        return None
    cursor.execute("SELECT id FROM genel_tanimlar WHERE grup=? AND deger=? AND firma_id=?", (grup, deger, firma_id))
    row = cursor.fetchone()
    if row:
        return row[0]
    return kaydet_kart(cursor, "genel_tanimlar", {"grup": grup, "deger": deger, "firma_id": firma_id})


def _grup_id_bul_veya_olustur(cursor, tur, grup_adi, firma_id):
    """Hizmet kartı grubunu bulur; yoksa oluşturur. Boşsa 'Diğer' kullanır."""
    grup_adi = grup_adi or "Diğer"
    cursor.execute(
        "SELECT id FROM hizmet_kartlari_gruplari WHERE firma_id=? AND tur=? AND grup_adi=?",
        (firma_id, tur, grup_adi),
    )
    row = cursor.fetchone()
    if row:
        return row[0]
    return kaydet_kart(
        cursor,
        "hizmet_kartlari_gruplari",
        {"grup_adi": grup_adi, "tur": tur, "firma_id": firma_id, "durum": 1},
    )


def tanim_import_dogrula(sayfalar, firma_id):
    """
    Excel'den okunan sayfaları doğrular.

    Dönüş: (hazir_kartlar, hatalar, uyarilar)
    """
    hatalar = []
    uyarilar = []
    hazir_kartlar = []

    conn = veritabani_baglan()
    try:
        cursor = conn.cursor()

        # Cari
        mevcut_cari_isimler = set()
        for row_no, raw in enumerate(sayfalar.get("Cari Kartlar", []), start=2):
            unvan = _metin(raw.get("Unvan"))
            if not unvan:
                hatalar.append(f"Cari Kartlar Satır {row_no}: Unvan boş bırakılamaz.")
                continue

            hata = _cari_duplicate_kontrol(cursor, unvan, firma_id, mevcut_cari_isimler)
            if hata:
                hatalar.append(f"Cari Kartlar Satır {row_no}: {hata}")
                continue

            tur = _metin(raw.get("Tür")) or "Müşteri"
            if tur not in ("Müşteri", "Tedarikçi", "Diğer"):
                hatalar.append(f"Cari Kartlar Satır {row_no}: Tür '{tur}' geçersiz.")
                continue

            mevcut_cari_isimler.add(unvan)
            hazir_kartlar.append({
                "tur": "Cari",
                "ad": unvan,
                "kod": "",
                "detay": f"{tur}",
                "tablo": "cariler",
                "veri": {
                    "unvan": unvan,
                    "tur": tur,
                    "telefon": _metin(raw.get("Telefon")),
                    "durum": _durum_to_int(raw.get("Durum")),
                    "firma_id": firma_id,
                },
            })

        # Stok
        mevcut_stok_adlar = set()
        mevcut_stok_kodlar = set()
        for row_no, raw in enumerate(sayfalar.get("Stok Kartları", []), start=2):
            stok_adi = _metin(raw.get("Stok Adı"))
            if not stok_adi:
                hatalar.append(f"Stok Kartları Satır {row_no}: Stok Adı boş bırakılamaz.")
                continue

            stok_kodu = _metin(raw.get("Stok Kodu"))
            hata = _stok_duplicate_kontrol(cursor, stok_adi, stok_kodu, firma_id, mevcut_stok_adlar, mevcut_stok_kodlar)
            if hata:
                hatalar.append(f"Stok Kartları Satır {row_no}: {hata}")
                continue

            kategori = _metin(raw.get("Kategori"))
            birim = _metin(raw.get("Birim"))

            alis = _sayi(raw.get("Alış Fiyatı")) or 0
            satis = _sayi(raw.get("Satış Fiyatı")) or 0
            kdv = _sayi(raw.get("KDV %"))
            if kdv is None:
                kdv = 20.0
            kritik = _sayi(raw.get("Kritik Miktar")) or 0

            mevcut_stok_adlar.add(stok_adi)
            if stok_kodu:
                mevcut_stok_kodlar.add(stok_kodu)

            hazir_kartlar.append({
                "tur": "Stok",
                "ad": stok_adi,
                "kod": stok_kodu,
                "detay": f"{birim or 'Adet'} / KDV %{kdv:g}",
                "tablo": "stoklar",
                "veri": {
                    "stok_adi": stok_adi,
                    "stok_kodu": stok_kodu,
                    "kategori": kategori,
                    "birim": birim or "Adet",
                    "alis_fiyati": alis,
                    "satis_fiyati": satis,
                    "kdv_oran": kdv,
                    "kritik_miktar": kritik,
                    "durum": _durum_to_int(raw.get("Durum")),
                    "firma_id": firma_id,
                },
                "yeni_kategori": kategori,
                "yeni_birim": birim,
            })

        # Hizmet
        mevcut_hizmet_adlar = set()
        for row_no, raw in enumerate(sayfalar.get("Hizmet Kartları", []), start=2):
            kart_adi = _metin(raw.get("Kart Adı"))
            if not kart_adi:
                hatalar.append(f"Hizmet Kartları Satır {row_no}: Kart Adı boş bırakılamaz.")
                continue

            hata = _hizmet_duplicate_kontrol(cursor, kart_adi, firma_id, mevcut_hizmet_adlar)
            if hata:
                hatalar.append(f"Hizmet Kartları Satır {row_no}: {hata}")
                continue

            tur = _metin(raw.get("Tür")) or "Gider"
            if tur not in ("Gider", "Gelir"):
                hatalar.append(f"Hizmet Kartları Satır {row_no}: Tür '{tur}' geçersiz.")
                continue

            grup_adi = _metin(raw.get("Grup"))
            kdv = _sayi(raw.get("KDV %"))
            if kdv is None:
                kdv = 20.0

            mevcut_hizmet_adlar.add(kart_adi)
            hazir_kartlar.append({
                "tur": "Hizmet",
                "ad": kart_adi,
                "kod": "",
                "detay": f"{tur} / KDV %{kdv:g}",
                "tablo": "hizmet_kartlari",
                "veri": {
                    "kart_adi": kart_adi,
                    "tur": tur,
                    "kdv_oran": kdv,
                    "durum": _durum_to_int(raw.get("Durum")),
                    "firma_id": firma_id,
                },
                "grup_adi": grup_adi,
            })

        # Kasa
        mevcut_kasa_adlar = set()
        for row_no, raw in enumerate(sayfalar.get("Kasa Kartları", []), start=2):
            kasa_adi = _metin(raw.get("Kasa Adı"))
            if not kasa_adi:
                hatalar.append(f"Kasa Kartları Satır {row_no}: Kasa Adı boş bırakılamaz.")
                continue

            hata = _kasa_duplicate_kontrol(cursor, kasa_adi, firma_id, mevcut_kasa_adlar)
            if hata:
                hatalar.append(f"Kasa Kartları Satır {row_no}: {hata}")
                continue

            mevcut_kasa_adlar.add(kasa_adi)
            hazir_kartlar.append({
                "tur": "Kasa",
                "ad": kasa_adi,
                "kod": "",
                "detay": "Kasa Kartı",
                "tablo": "kasalar",
                "veri": {
                    "kasa_adi": kasa_adi,
                    "durum": _durum_to_int(raw.get("Durum")),
                    "firma_id": firma_id,
                },
            })

        # Banka Kurumları
        mevcut_kurum_adlar = set()
        for row_no, raw in enumerate(sayfalar.get("Banka Kurumları", []), start=2):
            kurum_adi = _metin(raw.get("Kurum Adı"))
            if not kurum_adi:
                hatalar.append(f"Banka Kurumları Satır {row_no}: Kurum Adı boş bırakılamaz.")
                continue

            hata = _banka_kurum_duplicate_kontrol(cursor, kurum_adi, firma_id, mevcut_kurum_adlar)
            if hata:
                hatalar.append(f"Banka Kurumları Satır {row_no}: {hata}")
                continue

            mevcut_kurum_adlar.add(kurum_adi)
            hazir_kartlar.append({
                "tur": "Banka Kurum",
                "ad": kurum_adi,
                "kod": "",
                "detay": "Banka Kurumu",
                "tablo": "banka_kurumlari",
                "veri": {
                    "kurum_adi": kurum_adi,
                    "durum": _durum_to_int(raw.get("Durum")),
                    "firma_id": firma_id,
                },
            })

        # Banka Hesapları
        mevcut_hesap_adlar = set()
        for row_no, raw in enumerate(sayfalar.get("Banka Hesapları", []), start=2):
            hesap_adi = _metin(raw.get("Hesap Adı"))
            if not hesap_adi:
                hatalar.append(f"Banka Hesapları Satır {row_no}: Hesap Adı boş bırakılamaz.")
                continue

            hata = _banka_hesap_duplicate_kontrol(cursor, hesap_adi, firma_id, mevcut_hesap_adlar)
            if hata:
                hatalar.append(f"Banka Hesapları Satır {row_no}: {hata}")
                continue

            kurum_adi = _metin(raw.get("Banka Kurumu")) or "Diğer"
            hesap_turu = _metin(raw.get("Hesap Türü")) or "Vadesiz"
            if hesap_turu not in ("Vadesiz", "POS", "Kredi Kartı"):
                hatalar.append(f"Banka Hesapları Satır {row_no}: Hesap Türü '{hesap_turu}' geçersiz.")
                continue

            komisyon = _sayi(raw.get("Komisyon %"))
            if komisyon is None:
                komisyon = 0.0

            mevcut_hesap_adlar.add(hesap_adi)
            hazir_kartlar.append({
                "tur": "Banka Hesap",
                "ad": hesap_adi,
                "kod": kurum_adi,
                "detay": f"{kurum_adi} / {hesap_turu}",
                "tablo": "banka_hesaplari",
                "veri": {
                    "hesap_adi": hesap_adi,
                    "hesap_turu": hesap_turu,
                    "iban": _metin(raw.get("IBAN")),
                    "komisyon_orani": komisyon,
                    "durum": _durum_to_int(raw.get("Durum")),
                    "firma_id": firma_id,
                },
                "kurum_adi": kurum_adi,
            })

    finally:
        conn.close()

    return hazir_kartlar, hatalar, uyarilar


def tanim_verilerini_kaydet(cursor, hazir_kartlar, firma_id):
    """
    Doğrulanmış kart listesini veritabanına kaydeder.

    Stok kategori/birim ve hizmet grubu eksikse otomatik oluşturur.
    """
    eklenen_ids = []
    for kart in hazir_kartlar:
        veri = kart["veri"].copy()
        tablo = kart["tablo"]

        if tablo == "stoklar":
            if kart.get("yeni_kategori"):
                kat_id = _kategori_birim_id(cursor, "Stok Kategorisi", kart["yeni_kategori"], firma_id)
                if kat_id:
                    veri["kategori"] = kart["yeni_kategori"]
            if kart.get("yeni_birim"):
                birim_id = _kategori_birim_id(cursor, "Stok Birimi", kart["yeni_birim"], firma_id)
                if birim_id:
                    veri["birim"] = kart["yeni_birim"]

        elif tablo == "hizmet_kartlari":
            tur = veri["tur"]
            grup_id = _grup_id_bul_veya_olustur(cursor, tur, kart.get("grup_adi"), firma_id)
            veri["grup_id"] = grup_id

        elif tablo == "banka_hesaplari":
            veri["kurum_id"] = _kurum_id_bul_veya_olustur(cursor, kart.get("kurum_adi"), firma_id)

        eklenen_ids.append(kaydet_kart(cursor, tablo, veri))

    return eklenen_ids


class TanimImportMixin:
    """Tanım view'larına Örnek İndir / Veri Yükle buton davranışlarını ekler."""

    def ornek_indir(self):
        dosya_yolu = filedialog.asksaveasfilename(
            title="Örnek Tanım Excel Şablonunu Kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            initialfile="tanim_import_ornek.xlsx",
            parent=self,
        )
        if not dosya_yolu:
            return
        try:
            tanim_ornek_excel_olustur(dosya_yolu)
            messagebox.showinfo("Başarılı", "Örnek Excel dosyası oluşturuldu.\nCari, Stok, Hizmet, Kasa ve Banka sayfalarını doldurup Veri Yükle ile aktarabilirsiniz.", parent=self)
        except Exception as e:
            messagebox.showerror("Excel Oluşturma Hatası", f"Örnek dosya oluşturulamadı:\n{e}", parent=self)

    def veri_yukle(self):
        try:
            self._veri_yukle_impl()
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Veri Yükleme Hatası", f"Beklenmeyen bir hata oluştu:\n{e}", parent=self)

    def _veri_yukle_impl(self):
        dosya_yolu = filedialog.askopenfilename(
            title="Tanım Excel Dosyası Seç",
            filetypes=[("Excel Dosyası", "*.xlsx *.xls")],
            parent=self,
        )
        if not dosya_yolu:
            return

        sayfalar = tanim_excel_oku(dosya_yolu)
        hazir_kartlar, hatalar, uyarilar = tanim_import_dogrula(sayfalar, self.main_app.aktif_firma_id)

        if not hazir_kartlar and not hatalar:
            messagebox.showinfo("Bilgi", "Aktarılacak veri bulunamadı.\nExcel dosyası boş olabilir veya tanınan sayfa yok.", parent=self)
            return

        def import_callback():
            conn = None
            try:
                conn = veritabani_baglan()
                cursor = conn.cursor()
                eklenen_ids = tanim_verilerini_kaydet(cursor, hazir_kartlar, self.main_app.aktif_firma_id)
                conn.commit()
                mesaj = f"{len(eklenen_ids)} kart başarıyla içe aktarıldı."
                messagebox.showinfo("Başarılı", mesaj, parent=self)
                return True
            except Exception as e:
                if conn:
                    conn.rollback()
                messagebox.showerror("İçe Aktarma Hatası", f"Kartlar kaydedilirken bir hata oluştu:\n{e}", parent=self)
                return False
            finally:
                if conn:
                    conn.close()

        from ui.import_preview import DefinitionImportPreviewDialog

        try:
            DefinitionImportPreviewDialog(
                self,
                "Tanım İçe Aktarma Önizleme",
                hazir_kartlar,
                hatalar,
                uyarilar,
                on_import=import_callback,
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Önizleme Hatası", f"Önizleme ekranı açılırken bir hata oluştu:\n{e}", parent=self)

        if hasattr(self, "listele"):
            self.listele()
        elif hasattr(self, "yenile"):
            self.yenile()

class TanimImportView(TanimImportMixin, tk.Frame):
    """Tanımlar notebook'unun son sekmesinde gösterilen Excel import ekranı."""

    def __init__(self, parent, main_app):
        super().__init__(parent, bg="#f5f7fb")
        self.main_app = main_app
        self.parent = parent
        self.create_widgets()

    def create_widgets(self):
        ana_frame = tk.Frame(self, bg="#f5f7fb")
        ana_frame.pack(expand=True, fill="both", padx=40, pady=40)

        baslik = tk.Label(
            ana_frame,
            text="Excel Veri Yükleme",
            font=("Arial", 18, "bold"),
            bg="#f5f7fb",
            fg="#0d6efd",
        )
        baslik.pack(pady=(0, 10))

        aciklama = tk.Label(
            ana_frame,
            text=(
                "Tanım kartlarını tek Excel dosyasından içe aktarabilirsiniz.\n"
                "Cari, Stok, Hizmet, Kasa, Banka Kurumu ve Banka Hesapları sayfaları desteklenir."
            ),
            font=("Arial", 10),
            bg="#f5f7fb",
            justify="center",
        )
        aciklama.pack(pady=(0, 30))

        buton_frame = tk.Frame(ana_frame, bg="#f5f7fb")
        buton_frame.pack()

        tk.Button(
            buton_frame,
            text="Örnek İndir",
            command=self.ornek_indir,
            font=("Arial", 11, "bold"),
            padx=20,
            pady=8,
        ).pack(side="left", padx=10)

        tk.Button(
            buton_frame,
            text="Veri Yükle",
            command=self.veri_yukle,
            font=("Arial", 11, "bold"),
            bg="#198754",
            fg="white",
            padx=20,
            pady=8,
        ).pack(side="left", padx=10)

        bilgi = tk.Label(
            ana_frame,
            text=(
                "İpucu: Örnek dosyayı indirip sayfalardaki örnek satırları kendi verilerinizle değiştirin.\n"
                "Aynı isimde kart zaten tanımlıysa import hata verir."
            ),
            font=("Arial", 9),
            bg="#f5f7fb",
            fg="#6c757d",
            justify="center",
        )
        bilgi.pack(pady=(30, 0))
