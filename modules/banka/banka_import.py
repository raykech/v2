# -*- coding: utf-8 -*-
"""
Banka modülü Excel import yardımcıları.

- Örnek Excel şablonu oluşturur
- Excel dosyasını okur
- Satırları doğrular ve fiş/satır yapısına çevirir
- core.services.fis_kaydet ile kaydedilmeye hazır veri döndürür
"""
from datetime import datetime, date

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    load_workbook = None
    Alignment = Font = PatternFill = DataValidation = None

from core.db import veritabani_baglan
from core.services import fis_kaydet
from utils.formatters import kdv_hesapla


BANKA_FIS_TURLERI = [
    "Banka Gider Fişi",
    "Banka Gelir Fişi",
    "Bankalar Arası Virman",
    "Blokeyi Bankaya Aktar",
    "Bankaya Yatan",
    "Bankadan Çekilen",
    "Gelen Banka Transferi",
    "Giden Banka Transferi",
    "Banka Açılış Fişi",
]

TEK_TUTARLI_FIS_TURLERI = (
    "Bankalar Arası Virman",
    "Blokeyi Bankaya Aktar",
    "Bankaya Yatan",
    "Bankadan Çekilen",
    "Gelen Banka Transferi",
    "Giden Banka Transferi",
)

BANKA_IMPORT_KOLONLARI = [
    "Fiş Türü",
    "Tarih",
    "Fiş No",
    "Açıklama",
    "Ana Banka",
    "Hedef / Karşı Hesap",
    "Gider/Gelir Kartı",
    "Yön",
    "Satır Açıklaması",
    "Miktar",
    "Birim Fiyat",
    "KDV %",
    "Tutar",
]

ZORUNLU_KOLONLAR = ["Fiş Türü", "Tarih", "Ana Banka"]





from utils.import_helpers import metin as _metin, sayi as _sayi, tarih as _tarih, banka_id_bul as _banka_id_bul, kasa_id_bul as _kasa_id_bul, cari_id_bul as _cari_id_bul, hizmet_id_bul as _hizmet_id_bul

def banka_ornek_excel_olustur(dosya_yolu):
    """Banka importu için örnek Excel şablonu oluşturur."""
    if pd is None:
        raise RuntimeError("Excel şablonu oluşturmak için 'pandas' ve 'openpyxl' gereklidir.")

    ornek_satirlar = [
        # Banka Gider Fişi
        [
            "Banka Gider Fişi", "01.01.2026", "BG-001", "Vergi ödemesi", "Banka Hesabı 1",
            "", "Vergi Gideri", "", "Mart vergisi", 1, 5000, "", "",
        ],
        # Banka Gelir Fişi
        [
            "Banka Gelir Fişi", "02.01.2026", "BGL-001", "Tahsilat", "Banka Hesabı 1",
            "", "Satış Geliri", "", "", 1, 3000, "", "",
        ],
        # Bankalar Arası Virman
        [
            "Bankalar Arası Virman", "03.01.2026", "V-001", "Hesap aktarımı", "Banka Hesabı 1",
            "Banka Hesabı 2", "", "", "", "", "", "", 2000,
        ],
        # Bankaya Yatan
        [
            "Bankaya Yatan", "04.01.2026", "BY-001", "Nakit yatırma", "Banka Hesabı 1",
            "Ana Kasa", "", "", "", "", "", "", 1500,
        ],
        # Banka Açılış Fişi
        [
            "Banka Açılış Fişi", "01.01.2026", "A-001", "Açılış bakiyesi", "Banka Hesabı 1",
            "", "", "Borç", "Açılış", "", "", "", 10000,
        ],
    ]

    df = pd.DataFrame(ornek_satirlar, columns=BANKA_IMPORT_KOLONLARI)

    with pd.ExcelWriter(dosya_yolu, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Banka İşlemleri", index=False)

    if load_workbook is None:
        return

    wb = load_workbook(dosya_yolu)
    ws = wb["Banka İşlemleri"]

    dv = DataValidation(
        type="list",
        formula1='"Banka Gider Fişi,Banka Gelir Fişi,Bankalar Arası Virman,Blokeyi Bankaya Aktar,Bankaya Yatan,Bankadan Çekilen,Gelen Banka Transferi,Giden Banka Transferi,Banka Açılış Fişi"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add("A2:A2000")

    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_fill = PatternFill(start_color="4A69BD", end_color="4A69BD", fill_type="solid")
    for hucre in ws[1]:
        hucre.font = baslik_font
        hucre.fill = baslik_fill
        hucre.alignment = Alignment(horizontal="center", vertical="center")

    genislikler = {
        "A": 24, "B": 14, "C": 12, "D": 28, "E": 22,
        "F": 22, "G": 22, "H": 12, "I": 30, "J": 10,
        "K": 14, "L": 10, "M": 14,
    }
    for kolon, genislik in genislikler.items():
        ws.column_dimensions[kolon].width = genislik

    ws.freeze_panes = "A2"

    aciklama_ws = wb.create_sheet("Açıklama", 0)
    aciklama_ws.column_dimensions["A"].width = 120
    satirlar = [
        ("BANKA EXCEL İMPORT REHBERİ", True),
        ("", False),
        ("1. 'Banka İşlemleri' sayfasındaki örnek satırları kendi verilerinizle değiştirin.", False),
        ("2. Her satır bir fiş satırıdır. Aynı Fiş No'ya sahip satırlar tek fiş olarak gruplanır.", False),
        ("3. Fiş No boş bırakılırsa her Excel satırı ayrı bir fiş olarak içe aktarılır.", False),
        ("4. Gider/Gelir fişlerinde Miktar + Birim Fiyat veya sadece Tutar kullanılabilir.", False),
        ("5. Banka işlemleri KDV'sizdir; KDV % sütunu dikkate alınmaz.", False),
        ("6. Banka hesabı, kasa ve cari adları Tanımlar bölümündeki adlarla birebir aynı olmalıdır.", False),
        ("7. Blokeyi Bankaya Aktar fişinde Ana Banka POS, Hedef Banka normal banka hesabı olmalıdır.", False),
        ("8. Bankaya Yatan / Bankadan Çekilen fişlerinde karşı hesap Kasa'dır.", False),
        ("9. Gelen/Giden Banka Transferi fişlerinde karşı hesap Cari'dir.", False),
        ("10. Banka Açılış Fişinde Yön sütunu Borç veya Alacak olmalıdır.", False),
        ("", False),
        ("DESTEKLENEN FİŞ TÜRLERİ:", True),
        ("- Banka Gider Fişi", False),
        ("- Banka Gelir Fişi", False),
        ("- Bankalar Arası Virman", False),
        ("- Blokeyi Bankaya Aktar", False),
        ("- Bankaya Yatan", False),
        ("- Bankadan Çekilen", False),
        ("- Gelen Banka Transferi", False),
        ("- Giden Banka Transferi", False),
        ("- Banka Açılış Fişi", False),
    ]
    for i, (metin, kalin) in enumerate(satirlar, start=1):
        hucre = aciklama_ws.cell(row=i, column=1, value=metin)
        if kalin:
            hucre.font = Font(bold=True, size=12)

    wb.save(dosya_yolu)


def banka_excel_oku(dosya_yolu):
    """Excel dosyasındaki Banka İşlemleri sayfasını okur ve dict listesi döndürür."""
    if pd is None:
        raise RuntimeError("Excel okumak için 'pandas' ve 'openpyxl' gereklidir.")

    try:
        df = pd.read_excel(dosya_yolu, sheet_name="Banka İşlemleri")
    except Exception:
        sayfalar = pd.read_excel(dosya_yolu, sheet_name=None)
        df = None
        for _, sayfa in sayfalar.items():
            if all(kolon in sayfa.columns for kolon in ZORUNLU_KOLONLAR):
                df = sayfa
                break
        if df is None:
            raise ValueError("Excel dosyasında 'Banka İşlemleri' sayfası veya zorunlu sütunlar bulunamadı.")

    eksik = [kolon for kolon in ZORUNLU_KOLONLAR if kolon not in df.columns]
    if eksik:
        raise ValueError("Excel dosyasında şu zorunlu sütunlar bulunamadı: " + ", ".join(eksik))

    df = df.dropna(how="all")

    satirlar = []
    for _, row in df.iterrows():
        satir = {}
        for kolon in BANKA_IMPORT_KOLONLARI:
            satir[kolon] = row.get(kolon) if kolon in df.columns else ""
        satirlar.append(satir)
    return satirlar






def banka_import_dogrula(satirlar, firma_id, aktif_yil):
    """
    Excel satırlarını doğrular ve kaydedilmeye hazır fiş listesine çevirir.

    Dönüş: (hazir_fisler, hatalar, uyarilar)
    """
    hatalar = []
    uyarilar = []
    farkli_yillar = set()
    gruplar = {}

    conn = veritabani_baglan()
    try:
        cursor = conn.cursor()

        for index, raw_satir in enumerate(satirlar, start=2):
            row_no = index

            satir = {}
            for kolon in BANKA_IMPORT_KOLONLARI:
                deger = raw_satir.get(kolon, "") if isinstance(raw_satir, dict) else raw_satir
                if kolon in ("Miktar", "Birim Fiyat", "KDV %", "Tutar"):
                    satir[kolon] = _sayi(deger)
                elif kolon == "Tarih":
                    satir[kolon] = _tarih(deger)
                else:
                    satir[kolon] = _metin(deger)

            fis_turu = satir.get("Fiş Türü", "")
            if not fis_turu:
                hatalar.append(f"Satır {row_no}: Fiş Türü boş bırakılamaz.")
                continue
            if fis_turu not in BANKA_FIS_TURLERI:
                hatalar.append(f"Satır {row_no}: Geçersiz Fiş Türü '{fis_turu}'.")
                continue

            tarih = satir.get("Tarih")
            if not tarih:
                hatalar.append(f"Satır {row_no}: Tarih geçersiz veya boş.")
                continue

            if int(tarih[:4]) != aktif_yil:
                farkli_yillar.add(int(tarih[:4]))

            ana_banka_adi = satir.get("Ana Banka", "")
            if not ana_banka_adi:
                hatalar.append(f"Satır {row_no}: 'Ana Banka' boş bırakılamaz.")
                continue

            # Açılış fişinde ana banka satır bankasıdır; tür kısıtı yok.
            ana_banka_id = _banka_id_bul(cursor, ana_banka_adi, firma_id)
            if ana_banka_id is None:
                hatalar.append(f"Satır {row_no}: '{ana_banka_adi}' adında tanımlı banka hesabı bulunamadı.")
                continue
            if ana_banka_id == "ambiguous":
                hatalar.append(f"Satır {row_no}: '{ana_banka_adi}' adında birden fazla banka hesabı var.")
                continue

            fis_no = satir.get("Fiş No", "")
            genel_aciklama = satir.get("Açıklama", "")
            satir_aciklama = satir.get("Satır Açıklaması", "")

            # Banka Açılış Fişi
            if fis_turu == "Banka Açılış Fişi":
                yon = satir.get("Yön", "")
                if yon not in ("Borç", "Alacak"):
                    hatalar.append(f"Satır {row_no}: Yön 'Borç' veya 'Alacak' olmalıdır.")
                    continue

                tutar = satir.get("Tutar")
                if tutar is None or tutar <= 0:
                    hatalar.append(f"Satır {row_no}: Açılış tutarı 0'dan büyük olmalıdır.")
                    continue

                line = {
                    "hesap_turu": "Banka",
                    "hesap_id": ana_banka_id,
                    "borc": tutar if yon == "Borç" else 0,
                    "alacak": tutar if yon == "Alacak" else 0,
                    "aciklama": satir_aciklama,
                    "miktar": 1,
                    "birim_fiyat": tutar,
                    "kdv_oran": 0,
                    "kdv_tutar": 0,
                }
                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", None)
                _gruba_ekle(gruplar, anahtar, line, row_no, fis_turu, tarih, fis_no, genel_aciklama, ana_banka_adi, None, tutar)

            elif fis_turu in TEK_TUTARLI_FIS_TURLERI:
                hedef_adi = satir.get("Hedef / Karşı Hesap", "")
                if not hedef_adi:
                    hatalar.append(f"Satır {row_no}: 'Hedef / Karşı Hesap' boş bırakılamaz.")
                    continue

                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", None)
                if anahtar in gruplar:
                    hatalar.append(f"Satır {row_no}: Bu fiş türünde aynı Fiş No ile birden fazla satır bulunamaz.")
                    continue

                tutar = satir.get("Tutar")
                if tutar is None or tutar <= 0:
                    hatalar.append(f"Satır {row_no}: Tutar 0'dan büyük olmalıdır.")
                    continue

                if fis_turu == "Bankalar Arası Virman":
                    hedef_id = _banka_id_bul(cursor, hedef_adi, firma_id)
                    if hedef_id is None:
                        hatalar.append(f"Satır {row_no}: '{hedef_adi}' adında banka hesabı bulunamadı.")
                        continue
                    if hedef_id == "ambiguous":
                        hatalar.append(f"Satır {row_no}: '{hedef_adi}' adında birden fazla banka hesabı var.")
                        continue
                    if ana_banka_id == hedef_id:
                        hatalar.append(f"Satır {row_no}: Ana ve hedef banka hesapları aynı olamaz.")
                        continue

                    line1 = {
                        "hesap_turu": "Banka", "hesap_id": ana_banka_id,
                        "borc": 0, "alacak": tutar,
                        "aciklama": f"ID:{hedef_id} banka hesabına virman",
                        "miktar": 1, "birim_fiyat": tutar, "kdv_oran": 0, "kdv_tutar": 0,
                    }
                    line2 = {
                        "hesap_turu": "Banka", "hesap_id": hedef_id,
                        "borc": tutar, "alacak": 0,
                        "aciklama": f"ID:{ana_banka_id} banka hesabından virman",
                        "miktar": 1, "birim_fiyat": tutar, "kdv_oran": 0, "kdv_tutar": 0,
                    }

                elif fis_turu == "Blokeyi Bankaya Aktar":
                    pos_id = _banka_id_bul(cursor, ana_banka_adi, firma_id, hesap_turu="POS")
                    if pos_id is None or pos_id == "ambiguous":
                        hatalar.append(f"Satır {row_no}: Ana Banka POS hesabı olmalıdır.")
                        continue
                    hedef_id = _banka_id_bul(cursor, hedef_adi, firma_id)
                    if hedef_id is None:
                        hatalar.append(f"Satır {row_no}: '{hedef_adi}' adında banka hesabı bulunamadı.")
                        continue
                    if hedef_id == "ambiguous":
                        hatalar.append(f"Satır {row_no}: '{hedef_adi}' adında birden fazla banka hesabı var.")
                        continue

                    line1 = {
                        "hesap_turu": "Banka", "hesap_id": ana_banka_id,
                        "borc": 0, "alacak": tutar,
                        "aciklama": f"ID:{hedef_id} POS hesabından banka hesabına aktarım",
                        "miktar": 1, "birim_fiyat": tutar, "kdv_oran": 0, "kdv_tutar": 0,
                    }
                    line2 = {
                        "hesap_turu": "Banka", "hesap_id": hedef_id,
                        "borc": tutar, "alacak": 0,
                        "aciklama": f"ID:{ana_banka_id} POS hesabından aktarım",
                        "miktar": 1, "birim_fiyat": tutar, "kdv_oran": 0, "kdv_tutar": 0,
                    }

                elif fis_turu in ("Bankaya Yatan", "Bankadan Çekilen"):
                    normal_id = _banka_id_bul(cursor, ana_banka_adi, firma_id, hesap_turu="Vadesiz")
                    if normal_id is None or normal_id == "ambiguous":
                        hatalar.append(f"Satır {row_no}: Ana Banka normal (Vadesiz) banka hesabı olmalıdır.")
                        continue
                    karsi_id = _kasa_id_bul(cursor, hedef_adi, firma_id)
                    if karsi_id is None:
                        hatalar.append(f"Satır {row_no}: '{hedef_adi}' adında kasa bulunamadı.")
                        continue
                    if karsi_id == "ambiguous":
                        hatalar.append(f"Satır {row_no}: '{hedef_adi}' adında birden fazla kasa var.")
                        continue

                    banka_borclu = fis_turu == "Bankaya Yatan"
                    line1 = {
                        "hesap_turu": "Banka", "hesap_id": ana_banka_id,
                        "borc": tutar if banka_borclu else 0,
                        "alacak": 0 if banka_borclu else tutar,
                        "aciklama": f"{fis_turu} - karşı hesap ID:{karsi_id}",
                        "miktar": 1, "birim_fiyat": tutar, "kdv_oran": 0, "kdv_tutar": 0,
                    }
                    line2 = {
                        "hesap_turu": "Kasa", "hesap_id": karsi_id,
                        "borc": 0 if banka_borclu else tutar,
                        "alacak": tutar if banka_borclu else 0,
                        "aciklama": f"{fis_turu} - banka ID:{ana_banka_id}",
                        "miktar": 1, "birim_fiyat": tutar, "kdv_oran": 0, "kdv_tutar": 0,
                    }

                else:  # Gelen / Giden Banka Transferi
                    normal_id = _banka_id_bul(cursor, ana_banka_adi, firma_id, hesap_turu="Vadesiz")
                    if normal_id is None or normal_id == "ambiguous":
                        hatalar.append(f"Satır {row_no}: Ana Banka normal (Vadesiz) banka hesabı olmalıdır.")
                        continue
                    karsi_id = _cari_id_bul(cursor, hedef_adi, firma_id)
                    if karsi_id is None:
                        hatalar.append(f"Satır {row_no}: '{hedef_adi}' adında cari bulunamadı.")
                        continue
                    if karsi_id == "ambiguous":
                        hatalar.append(f"Satır {row_no}: '{hedef_adi}' adında birden fazla cari var.")
                        continue

                    banka_borclu = fis_turu == "Gelen Banka Transferi"
                    line1 = {
                        "hesap_turu": "Banka", "hesap_id": ana_banka_id,
                        "borc": tutar if banka_borclu else 0,
                        "alacak": 0 if banka_borclu else tutar,
                        "aciklama": f"{fis_turu} - cari ID:{karsi_id}",
                        "miktar": 1, "birim_fiyat": tutar, "kdv_oran": 0, "kdv_tutar": 0,
                    }
                    line2 = {
                        "hesap_turu": "Cari", "hesap_id": karsi_id,
                        "borc": 0 if banka_borclu else tutar,
                        "alacak": tutar if banka_borclu else 0,
                        "aciklama": f"{fis_turu} - banka ID:{ana_banka_id}",
                        "miktar": 1, "birim_fiyat": tutar, "kdv_oran": 0, "kdv_tutar": 0,
                    }

                _gruba_ekle(gruplar, anahtar, line1, row_no, fis_turu, tarih, fis_no, genel_aciklama, ana_banka_adi, hedef_adi, tutar)
                _gruba_ekle(gruplar, anahtar, line2, row_no, fis_turu, tarih, fis_no, genel_aciklama, ana_banka_adi, hedef_adi, tutar)
                gruplar[anahtar]["toplam_tutar"] = tutar
                gruplar[anahtar]["ana_banka_id"] = ana_banka_id

            else:  # Banka Gider / Gelir
                is_gider = fis_turu == "Banka Gider Fişi"
                beklenen_tur = "Gider" if is_gider else "Gelir"
                hesap_adi = satir.get("Gider/Gelir Kartı", "")
                if not hesap_adi:
                    hatalar.append(f"Satır {row_no}: Gider/Gelir kartı boş bırakılamaz.")
                    continue

                hesap_id = _hizmet_id_bul(cursor, hesap_adi, beklenen_tur, firma_id)
                if hesap_id is None:
                    hatalar.append(f"Satır {row_no}: '{hesap_adi}' adında uygun hizmet kartı bulunamadı.")
                    continue
                if hesap_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{hesap_adi}' adında birden fazla hizmet kartı var.")
                    continue
                if hesap_id == "wrong_type":
                    hatalar.append(f"Satır {row_no}: '{hesap_adi}' kartı bu fiş türü için uygun değil.")
                    continue

                miktar = satir.get("Miktar")
                birim_fiyat = satir.get("Birim Fiyat")
                tutar = satir.get("Tutar")
                # Banka gider/gelir fişleri KDV'sizdir; "KDV %" sütunu dikkate alınmaz.
                kdv_oran = 0.0

                if miktar is None:
                    miktar = 1.0
                if birim_fiyat is None and tutar is not None:
                    if miktar and miktar > 0:
                        birim_fiyat = tutar / miktar
                    else:
                        hatalar.append(f"Satır {row_no}: Miktar 0 olduğu için birim fiyat hesaplanamadı.")
                        continue
                if birim_fiyat is None:
                    hatalar.append(f"Satır {row_no}: Birim Fiyat veya Tutar girilmelidir.")
                    continue

                if miktar <= 0 or birim_fiyat <= 0:
                    hatalar.append(f"Satır {row_no}: Miktar ve Birim Fiyat 0'dan büyük olmalıdır.")
                    continue

                if tutar is not None and abs(miktar * birim_fiyat - tutar) > 0.01:
                    uyarilar.append(f"Satır {row_no}: Girilen Tutar ile Miktar*Birim Fiyat farklı. Miktar*Birim Fiyat esas alındı.")

                ara_toplam, kdv_tutar, genel_toplam = kdv_hesapla(miktar, birim_fiyat, kdv_oran)

                line = {
                    "hesap_turu": "Hizmet",
                    "hesap_id": hesap_id,
                    "borc": ara_toplam if is_gider else 0,
                    "alacak": 0 if is_gider else ara_toplam,
                    "aciklama": satir_aciklama,
                    "miktar": miktar,
                    "birim_fiyat": birim_fiyat,
                    "kdv_oran": kdv_oran,
                    "kdv_tutar": kdv_tutar,
                }
                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", ana_banka_id)
                _gruba_ekle(gruplar, anahtar, line, row_no, fis_turu, tarih, fis_no, genel_aciklama, ana_banka_adi, None, genel_toplam)
                gruplar[anahtar]["ana_banka_id"] = ana_banka_id

    finally:
        conn.close()

    hazir_fisler = []
    for grup in gruplar.values():
        if not grup["fis_satirlari"]:
            continue

        fis_turu = grup["fis_turu"]
        fis_satirlari = grup["fis_satirlari"]

        if fis_turu in ("Banka Gider Fişi", "Banka Gelir Fişi"):
            is_gider = fis_turu == "Banka Gider Fişi"
            toplam = grup["toplam_tutar"]
            karsi = {
                "hesap_turu": "Banka",
                "hesap_id": grup["ana_banka_id"],
                "borc": 0 if is_gider else toplam,
                "alacak": toplam if is_gider else 0,
                "aciklama": grup["genel_aciklama"],
                "miktar": 1,
                "birim_fiyat": toplam,
                "kdv_oran": 0,
                "kdv_tutar": 0,
            }
            fis_satirlari = fis_satirlari + [karsi]

        fis_baslik = {
            "tarih": grup["tarih"],
            "fis_turu": fis_turu,
            "fis_no": grup["fis_no"],
            "aciklama": grup["genel_aciklama"],
            "toplam_tutar": grup["toplam_tutar"],
            "cari_id": None,
            "firma_id": firma_id,
            "yil": int(grup["tarih"][:4]),
        }

        hazir_fisler.append({
            "fis_baslik": fis_baslik,
            "fis_satirlari": fis_satirlari,
            "satir_nos": grup["satir_nos"],
            "fis_turu": fis_turu,
            "tarih": grup["tarih"],
            "fis_no": grup["fis_no"],
            "kasa_adi": grup["kasa_adi"],
            "hedef_kasa_adi": grup.get("hedef_kasa_adi", ""),
            "toplam_tutar": grup["toplam_tutar"],
        })

    for farkli_yil in sorted(farkli_yillar):
        uyarilar.append(
            f"Seçili yıl ({aktif_yil}) dışında {farkli_yil} yılına ait satırlar bulundu. "
            f"Bu satırlar {farkli_yil} yılı olarak kaydedilecek."
        )

    return hazir_fisler, hatalar, uyarilar


def _gruba_ekle(gruplar, anahtar, line, row_no, fis_turu, tarih, fis_no, genel_aciklama, kasa_adi, hedef_kasa_adi, tutar):
    """Aynı fiş anahtarına satır ekler; grup yoksa oluşturur."""
    if anahtar not in gruplar:
        gruplar[anahtar] = {
            "fis_turu": fis_turu,
            "tarih": tarih,
            "fis_no": fis_no or "",
            "genel_aciklama": genel_aciklama,
            "ana_banka_id": None,
            "kasa_adi": kasa_adi,
            "hedef_kasa_adi": hedef_kasa_adi or "",
            "fis_satirlari": [],
            "satir_nos": [],
            "toplam_tutar": 0.0,
        }

    grup = gruplar[anahtar]
    if not grup["genel_aciklama"] and genel_aciklama:
        grup["genel_aciklama"] = genel_aciklama
    grup["fis_satirlari"].append(line)
    if row_no not in grup["satir_nos"]:
        grup["satir_nos"].append(row_no)
    grup["toplam_tutar"] = grup.get("toplam_tutar", 0.0) + tutar


def banka_fislerini_kaydet(cursor, hazir_fisler, firma_id, aktif_yil):
    """Doğrulanmış banka fiş listesini veritabanına kaydeder."""
    eklenen_ids = []
    for fis in hazir_fisler:
        yeni_id = fis_kaydet(cursor, fis["fis_baslik"], fis["fis_satirlari"], kaynak_modul="Banka")
        eklenen_ids.append(yeni_id)
    return eklenen_ids
