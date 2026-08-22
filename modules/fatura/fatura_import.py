# -*- coding: utf-8 -*-
"""
Fatura modülü Excel import yardımcıları.

- Örnek Excel şablonu oluşturur
- Excel dosyasını okur
- Satırları doğrular ve fatura/satır yapısına çevirir
- core.services.fis_kaydet ile kaydedilmeye hazır veri döndürür
"""
from datetime import datetime, date

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    load_workbook = None
    Alignment = Font = PatternFill = DataValidation = None

from core.db import veritabani_baglan
from core.services import fis_kaydet


FATURA_FIS_TURLERI = [
    "Satış Faturası",
    "Alış Faturası",
    "Satış İade Faturası",
    "Alış İade Faturası",
    "Hizmet Satış Faturası",
    "Hizmet Alış Faturası",
]

FATURA_IMPORT_KOLONLARI = [
    "Fiş Türü",
    "Tarih",
    "Fatura No",
    "Açıklama",
    "Cari",
    "Ödeme Tipi",
    "Ödeme Hesabı",
    "Stok/Hizmet Adı",
    "Satır Açıklaması",
    "Miktar",
    "Birim Fiyat",
    "KDV %",
    "Tutar",
]

ZORUNLU_KOLONLAR = ["Fiş Türü", "Tarih", "Stok/Hizmet Adı"]


def _metin(deger):
    if deger is None:
        return ""
    if isinstance(deger, float) and pd is not None and pd.isna(deger):
        return ""
    if isinstance(deger, str):
        return deger.strip()
    if isinstance(deger, (datetime, date)):
        return deger.strftime("%d.%m.%Y")
    return str(deger).strip()


def _sayi(deger):
    if deger is None:
        return None
    if isinstance(deger, bool):
        return None
    if isinstance(deger, (int, float)):
        if pd is not None and isinstance(deger, float) and pd.isna(deger):
            return None
        return float(deger)

    s = str(deger).strip().replace(" ", "").replace("TL", "").replace("₺", "").replace("%", "")
    if not s:
        return None

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    if s.count(".") > 1:
        s = s.replace(".", "")
    elif s.count(".") == 1 and len(s.split(".")[1]) == 3:
        s = s.replace(".", "")

    try:
        return float(s)
    except ValueError:
        return None


def _tarih(deger):
    if deger is None:
        return None
    if isinstance(deger, datetime):
        return deger.strftime("%Y-%m-%d")
    if isinstance(deger, date):
        return deger.strftime("%Y-%m-%d")

    s = str(deger).strip()
    if not s:
        return None

    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def fatura_ornek_excel_olustur(dosya_yolu):
    """Fatura importu için örnek Excel şablonu oluşturur."""
    if pd is None:
        raise RuntimeError("Excel şablonu oluşturmak için 'pandas' ve 'openpyxl' gereklidir.")

    ornek_satirlar = [
        # Satış faturası örneği
        [
            "Satış Faturası", "01.01.2026", "SF-001", "Ocak satışı", "Müşteri A",
            "Vadeli", "", "Ürün 1", "Satır açıklaması", 2, 100, 20, "",
        ],
        # Alış faturası örneği
        [
            "Alış Faturası", "02.01.2026", "AF-001", "Ocak alışı", "Tedarikçi B",
            "Banka", "Banka Hesabı 1", "Ürün 2", "", 5, 50, 20, "",
        ],
        # Hizmet satış faturası örneği
        [
            "Hizmet Satış Faturası", "03.01.2026", "HS-001", "Hizmet satışı", "Müşteri A",
            "Nakit", "Ana Kasa", "Danışmanlık", "", "", 1000, 20, "",
        ],
    ]

    df = pd.DataFrame(ornek_satirlar, columns=FATURA_IMPORT_KOLONLARI)

    with pd.ExcelWriter(dosya_yolu, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Fatura İşlemleri", index=False)

    if load_workbook is None:
        return

    wb = load_workbook(dosya_yolu)
    ws = wb["Fatura İşlemleri"]

    dv = DataValidation(
        type="list",
        formula1='"Satış Faturası,Alış Faturası,Satış İade Faturası,Alış İade Faturası,Hizmet Satış Faturası,Hizmet Alış Faturası"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add("A2:A2000")

    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_fill = PatternFill(start_color="4A69BD", end_color="4A69BD", fill_type="solid")
    for hucre in ws[1]:
        hucre.font = baslik_font
        hucre.fill = baslik_fill
        hucre.alignment = Alignment(horizontal="center", vertical="center")

    genislikler = {
        "A": 24, "B": 14, "C": 14, "D": 28, "E": 22,
        "F": 14, "G": 22, "H": 22, "I": 30, "J": 10,
        "K": 14, "L": 10, "M": 14,
    }
    for kolon, genislik in genislikler.items():
        ws.column_dimensions[kolon].width = genislik

    ws.freeze_panes = "A2"

    aciklama_ws = wb.create_sheet("Açıklama", 0)
    aciklama_ws.column_dimensions["A"].width = 120
    satirlar = [
        ("FATURA EXCEL İMPORT REHBERİ", True),
        ("", False),
        ("1. 'Fatura İşlemleri' sayfasındaki örnek satırları kendi verilerinizle değiştirin.", False),
        ("2. Her satır bir fatura satırıdır. Aynı Fatura No'ya sahip satırlar tek fatura olarak gruplanır.", False),
        ("3. Fatura No boş bırakılırsa her Excel satırı ayrı bir fatura olarak içe aktarılır.", False),
        ("4. Stoklu faturalarda Miktar + Birim Fiyat, hizmet faturalarında Birim Fiyat/Tutar kullanılır.", False),
        ("5. KDV % boş bırakılırsa 0 kabul edilir.", False),
        ("6. Vadeli faturalarda Cari zorunludur, Ödeme Hesabı boş olur.", False),
        ("7. Nakit/Banka/POS faturalarında Ödeme Hesabı zorunludur.", False),
        ("8. Cari, Stok, Hizmet, Kasa ve Banka adları Tanımlar bölümüyle birebir aynı olmalıdır.", False),
        ("", False),
        ("DESTEKLENEN FİŞ TÜRLERİ:", True),
        ("- Satış Faturası", False),
        ("- Alış Faturası", False),
        ("- Satış İade Faturası", False),
        ("- Alış İade Faturası", False),
        ("- Hizmet Satış Faturası", False),
        ("- Hizmet Alış Faturası", False),
    ]
    for i, (metin, kalin) in enumerate(satirlar, start=1):
        hucre = aciklama_ws.cell(row=i, column=1, value=metin)
        if kalin:
            hucre.font = Font(bold=True, size=12)

    wb.save(dosya_yolu)


def fatura_excel_oku(dosya_yolu):
    """Excel dosyasındaki Fatura İşlemleri sayfasını okur ve dict listesi döndürür."""
    if pd is None:
        raise RuntimeError("Excel okumak için 'pandas' ve 'openpyxl' gereklidir.")

    try:
        df = pd.read_excel(dosya_yolu, sheet_name="Fatura İşlemleri")
    except Exception:
        sayfalar = pd.read_excel(dosya_yolu, sheet_name=None)
        df = None
        for _, sayfa in sayfalar.items():
            if all(kolon in sayfa.columns for kolon in ZORUNLU_KOLONLAR):
                df = sayfa
                break
        if df is None:
            raise ValueError("Excel dosyasında 'Fatura İşlemleri' sayfası veya zorunlu sütunlar bulunamadı.")

    eksik = [kolon for kolon in ZORUNLU_KOLONLAR if kolon not in df.columns]
    if eksik:
        raise ValueError("Excel dosyasında şu zorunlu sütunlar bulunamadı: " + ", ".join(eksik))

    df = df.dropna(how="all")

    satirlar = []
    for _, row in df.iterrows():
        satir = {}
        for kolon in FATURA_IMPORT_KOLONLARI:
            satir[kolon] = row.get(kolon) if kolon in df.columns else ""
        satirlar.append(satir)
    return satirlar


def _cari_id_bul(cursor, unvan, firma_id):
    if not unvan:
        return None
    cursor.execute("SELECT id FROM cariler WHERE durum=1 AND firma_id=? AND unvan=?", (firma_id, unvan))
    kayitlar = cursor.fetchall()
    if not kayitlar:
        return None
    if len(kayitlar) > 1:
        return "ambiguous"
    return kayitlar[0][0]


def _stok_id_bul(cursor, stok_adi, firma_id):
    if not stok_adi:
        return None
    cursor.execute("SELECT id, kdv_oran FROM stoklar WHERE durum=1 AND firma_id=? AND stok_adi=?", (firma_id, stok_adi))
    kayitlar = cursor.fetchall()
    if not kayitlar:
        return None
    if len(kayitlar) > 1:
        return "ambiguous"
    return kayitlar[0][0]


def _hizmet_id_bul(cursor, kart_adi, tur, firma_id):
    if not kart_adi:
        return None
    cursor.execute("SELECT id, tur, kdv_oran FROM hizmet_kartlari WHERE durum=1 AND firma_id=? AND kart_adi=?", (firma_id, kart_adi))
    kayitlar = cursor.fetchall()
    if not kayitlar:
        return None
    if len(kayitlar) > 1:
        return "ambiguous"
    if kayitlar[0][1] != tur:
        return "wrong_type"
    return kayitlar[0][0]


def _odeme_hesap_id_bul(cursor, hesap_adi, hesap_turu, firma_id):
    if not hesap_adi:
        return None
    if hesap_turu == "Nakit":
        cursor.execute("SELECT id FROM kasalar WHERE durum=1 AND firma_id=? AND kasa_adi=?", (firma_id, hesap_adi))
    else:
        banka_hesap_turu = "Vadesiz" if hesap_turu == "Banka" else "POS"
        cursor.execute("SELECT id FROM banka_hesaplari WHERE durum=1 AND firma_id=? AND hesap_adi=? AND hesap_turu=?", (firma_id, hesap_adi, banka_hesap_turu))
    kayitlar = cursor.fetchall()
    if not kayitlar:
        return None
    if len(kayitlar) > 1:
        return "ambiguous"
    return kayitlar[0][0]


def fatura_import_dogrula(satirlar, firma_id, aktif_yil):
    """
    Excel satırlarını doğrular ve kaydedilmeye hazır fatura listesine çevirir.

    Dönüş: (hazir_faturalar, hatalar, uyarilar)
    """
    hatalar = []
    uyarilar = []
    gruplar = {}

    conn = veritabani_baglan()
    try:
        cursor = conn.cursor()

        for index, raw_satir in enumerate(satirlar, start=2):
            row_no = index

            satir = {}
            for kolon in FATURA_IMPORT_KOLONLARI:
                deger = raw_satir.get(kolon, "") if isinstance(raw_satir, dict) else raw_satir
                if kolon in ("Miktar", "Birim Fiyat", "KDV %", "Tutar"):
                    satir[kolon] = _sayi(deger)
                elif kolon == "Tarih":
                    satir[kolon] = _tarih(deger)
                else:
                    satir[kolon] = _metin(deger)

            fis_turu = satir.get("Fiş Türü", "")
            if not fis_turu:
                hatalar.append(f"Satır {row_no}: Fiş Türü boş bırakılamaz.")
                continue
            if fis_turu not in FATURA_FIS_TURLERI:
                hatalar.append(f"Satır {row_no}: Geçersiz Fiş Türü '{fis_turu}'.")
                continue

            tarih = satir.get("Tarih")
            if not tarih:
                hatalar.append(f"Satır {row_no}: Tarih geçersiz veya boş.")
                continue

            fis_no = satir.get("Fatura No", "")
            genel_aciklama = satir.get("Açıklama", "")
            satir_aciklama = satir.get("Satır Açıklaması", "")
            odeme_tipi = satir.get("Ödeme Tipi", "") or "Vadeli"
            if odeme_tipi not in ("Vadeli", "Nakit", "Banka", "POS"):
                hatalar.append(f"Satır {row_no}: Ödeme Tipi '{odeme_tipi}' geçersiz.")
                continue

            cari_adi = satir.get("Cari", "")
            cari_id = None
            if odeme_tipi == "Vadeli":
                if not cari_adi:
                    hatalar.append(f"Satır {row_no}: Vadeli fatura için Cari zorunludur.")
                    continue
                cari_id = _cari_id_bul(cursor, cari_adi, firma_id)
                if cari_id is None:
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında tanımlı cari bulunamadı.")
                    continue
                if cari_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında birden fazla cari var.")
                    continue

            odeme_hesap_adi = satir.get("Ödeme Hesabı", "")
            odeme_hesap_id = None
            if odeme_tipi != "Vadeli":
                if not odeme_hesap_adi:
                    hatalar.append(f"Satır {row_no}: {odeme_tipi} ödeme tipi için Ödeme Hesabı zorunludur.")
                    continue
                odeme_hesap_id = _odeme_hesap_id_bul(cursor, odeme_hesap_adi, odeme_tipi, firma_id)
                if odeme_hesap_id is None:
                    hatalar.append(f"Satır {row_no}: '{odeme_hesap_adi}' adında uygun {odeme_tipi} hesabı bulunamadı.")
                    continue
                if odeme_hesap_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{odeme_hesap_adi}' adında birden fazla uygun hesap var.")
                    continue

            is_hizmet = "Hizmet" in fis_turu
            hesap_adi = satir.get("Stok/Hizmet Adı", "")
            if not hesap_adi:
                hatalar.append(f"Satır {row_no}: Stok/Hizmet Adı boş bırakılamaz.")
                continue

            if is_hizmet:
                beklenen_tur = "Gelir" if "Satış" in fis_turu else "Gider"
                hesap_id = _hizmet_id_bul(cursor, hesap_adi, beklenen_tur, firma_id)
                if hesap_id is None:
                    hatalar.append(f"Satır {row_no}: '{hesap_adi}' adında uygun hizmet kartı bulunamadı.")
                    continue
                if hesap_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{hesap_adi}' adında birden fazla hizmet kartı var.")
                    continue
                if hesap_id == "wrong_type":
                    hatalar.append(f"Satır {row_no}: '{hesap_adi}' kartı bu fatura türü için uygun değil.")
                    continue
            else:
                hesap_id = _stok_id_bul(cursor, hesap_adi, firma_id)
                if hesap_id is None:
                    hatalar.append(f"Satır {row_no}: '{hesap_adi}' adında tanımlı stok bulunamadı.")
                    continue
                if hesap_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{hesap_adi}' adında birden fazla stok var.")
                    continue

            miktar = satir.get("Miktar")
            birim_fiyat = satir.get("Birim Fiyat")
            tutar = satir.get("Tutar")
            kdv_oran = satir.get("KDV %")
            if kdv_oran is None:
                kdv_oran = 0.0

            if is_hizmet:
                miktar = 1.0
                if birim_fiyat is None and tutar is not None:
                    birim_fiyat = tutar
            else:
                if miktar is None:
                    miktar = 1.0
                if birim_fiyat is None and tutar is not None:
                    if miktar and miktar > 0:
                        birim_fiyat = tutar / miktar
                    else:
                        hatalar.append(f"Satır {row_no}: Miktar 0 olduğu için birim fiyat hesaplanamadı.")
                        continue

            if birim_fiyat is None:
                hatalar.append(f"Satır {row_no}: Birim Fiyat veya Tutar girilmelidir.")
                continue

            if miktar <= 0 or birim_fiyat <= 0:
                hatalar.append(f"Satır {row_no}: Miktar ve Birim Fiyat 0'dan büyük olmalıdır.")
                continue

            if tutar is not None and abs(miktar * birim_fiyat - tutar) > 0.01:
                uyarilar.append(f"Satır {row_no}: Girilen Tutar ile Miktar*Birim Fiyat farklı. Miktar*Birim Fiyat esas alındı.")

            ara_toplam = miktar * birim_fiyat
            kdv_tutar = ara_toplam * (kdv_oran / 100.0)
            genel_toplam = ara_toplam + kdv_tutar

            # Fatura satır yönü
            if "Satış Faturası" in fis_turu or "Alış İade Faturası" in fis_turu:
                borc, alacak = 0.0, genel_toplam
            else:
                borc, alacak = genel_toplam, 0.0

            line = {
                "hesap_turu": "Hizmet" if is_hizmet else "Stok",
                "hesap_id": hesap_id,
                "borc": borc,
                "alacak": alacak,
                "aciklama": satir_aciklama or f"{hesap_adi} - {fis_turu}",
                "miktar": miktar,
                "birim_fiyat": birim_fiyat,
                "kdv_oran": kdv_oran,
                "kdv_tutar": kdv_tutar,
            }

            # Grup anahtarı: aynı fatura no + tarih + tür + cari + ödeme
            anahtar = (
                fis_turu,
                tarih,
                fis_no or f"#SATIR{row_no}",
                cari_adi,
                odeme_tipi,
                odeme_hesap_adi,
            )
            if anahtar not in gruplar:
                gruplar[anahtar] = {
                    "fis_turu": fis_turu,
                    "tarih": tarih,
                    "fis_no": fis_no or "",
                    "genel_aciklama": genel_aciklama,
                    "cari_adi": cari_adi,
                    "cari_id": cari_id,
                    "odeme_tipi": odeme_tipi,
                    "odeme_hesap_adi": odeme_hesap_adi,
                    "odeme_hesap_id": odeme_hesap_id,
                    "fis_satirlari": [],
                    "satir_nos": [],
                    "toplam_tutar": 0.0,
                }

            grup = gruplar[anahtar]
            if not grup["genel_aciklama"] and genel_aciklama:
                grup["genel_aciklama"] = genel_aciklama
            grup["fis_satirlari"].append(line)
            if row_no not in grup["satir_nos"]:
                grup["satir_nos"].append(row_no)
            grup["toplam_tutar"] = grup.get("toplam_tutar", 0.0) + genel_toplam

    finally:
        conn.close()

    hazir_faturalar = []
    for grup in gruplar.values():
        if not grup["fis_satirlari"]:
            continue

        fis_turu = grup["fis_turu"]
        fis_satirlari = grup["fis_satirlari"]
        toplam = grup["toplam_tutar"]
        cari_id = grup["cari_id"]

        # Vadeli faturada cari karşılık satırı
        if grup["odeme_tipi"] == "Vadeli" and cari_id:
            is_satis = ("Satış Faturası" in fis_turu) or ("Hizmet Satış" in fis_turu)
            is_iade = "İade" in fis_turu
            cari_borclu = (is_satis and not is_iade) or (not is_satis and is_iade)
            fis_satirlari = fis_satirlari + [{
                "hesap_turu": "Cari",
                "hesap_id": cari_id,
                "borc": toplam if cari_borclu else 0,
                "alacak": 0 if cari_borclu else toplam,
                "aciklama": f"{fis_turu} cari karşılığı",
                "miktar": None,
                "birim_fiyat": None,
                "kdv_oran": None,
                "kdv_tutar": None,
            }]

        pesin_odeme_data = None
        if grup["odeme_tipi"] != "Vadeli":
            odeme_tipi = grup["odeme_tipi"]
            odeme_hesap_id = grup["odeme_hesap_id"]
            odeme_hesap_turu = {"Nakit": "Kasa", "Banka": "Banka", "POS": "Banka"}[odeme_tipi]
            is_tahsilat = "Satış Faturası" in fis_turu or "Alış İade Faturası" in fis_turu
            odeme_fis_turu = f"Fatura Peşin Tahsilat ({odeme_tipi})" if is_tahsilat else f"Fatura Peşin Ödeme ({odeme_tipi})"

            pesin_odeme_data = {
                "tarih": grup["tarih"],
                "fis_turu": odeme_fis_turu,
                "toplam_tutar": toplam,
                "kaynak_modul": "Fatura",
                "aciklama": f"Fatura No: {grup['fis_no']} peşin ödemesi",
                "firma_id": firma_id,
                "yil": int(tarih[:4]),
            }
            pesin_odeme_data["satirlar"] = list(grup["fis_satirlari"])
            pesin_odeme_data["satirlar"].append({
                "hesap_turu": odeme_hesap_turu,
                "hesap_id": odeme_hesap_id,
                "borc": toplam if is_tahsilat else 0,
                "alacak": 0 if is_tahsilat else toplam,
                "aciklama": f"{fis_turu} peşin ödemesi",
            })

        fis_baslik = {
            "tarih": grup["tarih"],
            "fis_turu": fis_turu,
            "fis_no": grup["fis_no"],
            "aciklama": grup["genel_aciklama"],
            "cari_id": cari_id if grup["odeme_tipi"] == "Vadeli" else None,
            "toplam_tutar": toplam,
            "firma_id": firma_id,
            "yil": int(tarih[:4]),
        }

        hazir_faturalar.append({
            "fis_baslik": fis_baslik,
            "fis_satirlari": fis_satirlari,
            "pesin_odeme_data": pesin_odeme_data,
            "satir_nos": grup["satir_nos"],
            "fis_turu": fis_turu,
            "tarih": grup["tarih"],
            "fis_no": grup["fis_no"],
            "cari_adi": grup["cari_adi"],
            "odeme_tipi": grup["odeme_tipi"],
            "odeme_hesap_adi": grup["odeme_hesap_adi"],
            "toplam_tutar": toplam,
        })

    return hazir_faturalar, hatalar, uyarilar


def fatura_fislerini_kaydet(cursor, hazir_faturalar, firma_id, aktif_yil):
    """Doğrulanmış fatura listesini veritabanına kaydeder ve ID'leri döndürür."""
    eklenen_ids = []
    for fatura in hazir_faturalar:
        yeni_id = fis_kaydet(
            cursor,
            fatura["fis_baslik"],
            fatura["fis_satirlari"],
            pesin_odeme_data=fatura["pesin_odeme_data"],
            kaynak_modul="Fatura",
        )
        eklenen_ids.append(yeni_id)
    return eklenen_ids
