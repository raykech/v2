# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, date

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    load_workbook = None
    Alignment = Font = PatternFill = DataValidation = None

from core.db import veritabani_baglan
from core.services import fis_kaydet, cek_senet_hareket_ekle


CEK_SENET_FIS_TURLERI = [
    "Çek/Senet Giriş Fişi",
    "Çek/Senet Açılış Fişi",
    "Çek/Senet Bankaya Tahsile Verme",
    "Çek/Senet Ciro Etme",
    "Çek/Senet Tahsil Fişi",
    "Çek/Senet İade Fişi",
]

CEK_SENET_IMPORT_KOLONLARI = [
    "Fiş Türü",
    "Tarih",
    "Fiş No",
    "Açıklama",
    "Seri No",
    "Tür",
    "Banka Kurumu",
    "Vade",
    "Tutar",
    "Keşideci",
    "Ciranta",
    "Cari / Karşı Hesap",
    "Satır Açıklaması",
    "Kasa / Banka Hesabı",
    "Tahsil Türü",
    "Durum",
]

ZORUNLU_KOLONLAR = ["Fiş Türü", "Tarih", "Seri No", "Tutar"]





from utils.import_helpers import metin as _metin, sayi as _sayi, tarih as _tarih, cari_id_bul as _cari_id_bul, banka_kurum_id_bul as _banka_kurum_id_bul

def cek_senet_ornek_excel_olustur(dosya_yolu):
    if pd is None:
        raise RuntimeError("Excel şablonu oluşturmak için 'pandas' ve 'openpyxl' gereklidir.")
    ornek = [
        ["Çek/Senet Giriş Fişi", "01.01.2026", "GS-001", "Müşteriden alınan çek", "C-001", "Çek", "Banka A", "30.06.2026", 5000, "Firma X", "", "Müşteri A", "", "", "", "Portföyde"],
        ["Çek/Senet Açılış Fişi", "01.01.2026", "AS-001", "Açılış", "C-002", "Senet", "Banka A", "31.12.2026", 10000, "Firma Y", "", "", "", "", "", "Portföyde"],
    ]
    df = pd.DataFrame(ornek, columns=CEK_SENET_IMPORT_KOLONLARI)
    with pd.ExcelWriter(dosya_yolu, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Çek/Senet İşlemleri", index=False)
    if load_workbook is None: return
    wb = load_workbook(dosya_yolu)
    ws = wb["Çek/Senet İşlemleri"]
    dv = DataValidation(type="list", formula1='"Çek/Senet Giriş Fişi,Çek/Senet Açılış Fişi"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv); dv.add("A2:A2000")
    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_fill = PatternFill(start_color="4A69BD", end_color="4A69BD", fill_type="solid")
    for hucre in ws[1]:
        hucre.font = baslik_font; hucre.fill = baslik_fill; hucre.alignment = Alignment(horizontal="center", vertical="center")
    for c, w in [("A", 26), ("B", 14), ("C", 12), ("D", 28), ("E", 16), ("F", 10), ("G", 18), ("H", 14), ("I", 14), ("J", 16), ("K", 16), ("L", 22), ("M", 30), ("N", 22), ("O", 14), ("P", 14)]:
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A2"
    aciklama_ws = wb.create_sheet("Açıklama", 0)
    aciklama_ws.column_dimensions["A"].width = 120
    satirlar = [
        ("ÇEK/SENET EXCEL İMPORT REHBERİ", True),
        ("", False),
        ("Yalnızca Giriş Fişi ve Açılış Fişi desteklenir. Diğer fiş türleri için manuel form kullanın.", False),
        ("Seri No benzersiz olmalıdır; aynı seri no varsa hata verilir.", False),
        ("Cari / Karşı Hesap: Giriş fişinde müşteri/cari adı.", False),
        ("Kasa / Banka Hesabı ve Tahsil Türü diğer fiş türleri için kullanılır, bu şablonda boş bırakılır.", False),
    ]
    for i, (metin, kalin) in enumerate(satirlar, start=1):
        hucre = aciklama_ws.cell(row=i, column=1, value=metin)
        if kalin: hucre.font = Font(bold=True, size=12)
    wb.save(dosya_yolu)


def cek_senet_excel_oku(dosya_yolu):
    if pd is None:
        raise RuntimeError("Excel okumak için 'pandas' ve 'openpyxl' gereklidir.")
    try:
        df = pd.read_excel(dosya_yolu, sheet_name="Çek/Senet İşlemleri")
    except Exception:
        sayfalar = pd.read_excel(dosya_yolu, sheet_name=None)
        df = None
        for _, sayfa in sayfalar.items():
            if all(kolon in sayfa.columns for kolon in ZORUNLU_KOLONLAR):
                df = sayfa; break
        if df is None:
            raise ValueError("Excel dosyasında 'Çek/Senet İşlemleri' sayfası veya zorunlu sütunlar bulunamadı.")
    eksik = [kolon for kolon in ZORUNLU_KOLONLAR if kolon not in df.columns]
    if eksik:
        raise ValueError("Excel dosyasında şu zorunlu sütunlar bulunamadı: " + ", ".join(eksik))
    df = df.dropna(how="all")
    satirlar = []
    for _, row in df.iterrows():
        satir = {}
        for kolon in CEK_SENET_IMPORT_KOLONLARI:
            satir[kolon] = row.get(kolon) if kolon in df.columns else ""
        satirlar.append(satir)
    return satirlar




def _cek_senet_seri_no_kontrol(cursor, seri_no, firma_id):
    cursor.execute("SELECT id FROM cekler_senetler WHERE firma_id=? AND seri_no=?", (firma_id, seri_no))
    return cursor.fetchone() is not None


def cek_senet_import_dogrula(satirlar, firma_id, aktif_yil):
    hatalar = []
    uyarilar = []
    farkli_yillar = set()
    hazir_fisler = []

    conn = veritabani_baglan()
    try:
        cursor = conn.cursor()

        for index, raw_satir in enumerate(satirlar, start=2):
            row_no = index
            satir = {}
            for kolon in CEK_SENET_IMPORT_KOLONLARI:
                deger = raw_satir.get(kolon, "") if isinstance(raw_satir, dict) else raw_satir
                if kolon == "Tutar":
                    satir[kolon] = _sayi(deger)
                elif kolon in ("Tarih", "Vade"):
                    satir[kolon] = _tarih(deger)
                else:
                    satir[kolon] = _metin(deger)

            fis_turu = satir.get("Fiş Türü", "")
            if not fis_turu:
                hatalar.append(f"Satır {row_no}: Fiş Türü boş.")
                continue
            if fis_turu not in ("Çek/Senet Giriş Fişi", "Çek/Senet Açılış Fişi"):
                hatalar.append(f"Satır {row_no}: '{fis_turu}' henüz import desteklemiyor. Yalnızca Giriş ve Açılış fişleri desteklenir.")
                continue

            tarih = satir.get("Tarih")
            if not tarih:
                hatalar.append(f"Satır {row_no}: Tarih geçersiz.")
                continue

            if int(tarih[:4]) != aktif_yil:
                farkli_yillar.add(int(tarih[:4]))

            seri_no = satir.get("Seri No", "")
            if not seri_no:
                hatalar.append(f"Satır {row_no}: Seri No boş.")
                continue
            if _cek_senet_seri_no_kontrol(cursor, seri_no, firma_id):
                hatalar.append(f"Satır {row_no}: '{seri_no}' seri no zaten kullanılmış.")
                continue

            turu = satir.get("Tür", "") or "Çek"
            if turu not in ("Çek", "Senet"):
                hatalar.append(f"Satır {row_no}: Tür 'Çek' veya 'Senet' olmalıdır.")
                continue

            vade = satir.get("Vade")
            if not vade:
                hatalar.append(f"Satır {row_no}: Vade tarihi boş.")
                continue

            tutar = satir.get("Tutar")
            if tutar is None or tutar <= 0:
                hatalar.append(f"Satır {row_no}: Tutar 0'dan büyük olmalıdır.")
                continue

            banka_kurum_adi = satir.get("Banka Kurumu", "")
            banka_kurum_id = None
            if banka_kurum_adi:
                banka_kurum_id = _banka_kurum_id_bul(cursor, banka_kurum_adi, firma_id)
                if banka_kurum_id is None:
                    hatalar.append(f"Satır {row_no}: '{banka_kurum_adi}' adında banka kurumu bulunamadı.")
                    continue
                if banka_kurum_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{banka_kurum_adi}' adında birden fazla banka kurumu var.")
                    continue

            kesideci = satir.get("Keşideci", "")
            ciranta = satir.get("Ciranta", "")
            satir_aciklama = satir.get("Satır Açıklaması", "")
            durum = satir.get("Durum", "") or "Portföyde"
            if durum not in ("Portföyde", "Bankada Tahsilde", "Cirolu", "Tahsil Edildi", "İade Edildi"):
                hatalar.append(f"Satır {row_no}: Geçersiz durum '{durum}'.")
                continue

            # Giriş fişinde cari
            cari_id = None
            cari_adi = None
            if fis_turu == "Çek/Senet Giriş Fişi":
                cari_adi = satir.get("Cari / Karşı Hesap", "")
                if not cari_adi:
                    hatalar.append(f"Satır {row_no}: Giriş fişi için Cari zorunludur.")
                    continue
                cari_id = _cari_id_bul(cursor, cari_adi, firma_id)
                if cari_id is None:
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında cari bulunamadı.")
                    continue
                if cari_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında birden fazla cari var.")
                    continue

            # Çek/Senet kartı verisi
            cek_senet_data = {
                "seri_no": seri_no,
                "turu": turu,
                "banka_id": banka_kurum_id,
                "vade": vade,
                "tutar": tutar,
                "kesideci": kesideci,
                "ciranta": ciranta,
                "aciklama": satir_aciklama,
                "durum": durum,
            }

            # Fis satırları
            fis_satirlari = [{
                "hesap_turu": "CekSenet",
                "hesap_id": 0,  # placeholder
                "borc": tutar,
                "alacak": 0,
                "aciklama": satir_aciklama,
                "miktar": 1,
                "birim_fiyat": tutar,
                "kdv_oran": 0,
                "kdv_tutar": 0,
            }]
            if cari_id:
                fis_satirlari.append({
                    "hesap_turu": "Cari",
                    "hesap_id": cari_id,
                    "borc": 0,
                    "alacak": tutar,
                    "aciklama": satir_aciklama,
                    "miktar": 1,
                    "birim_fiyat": tutar,
                    "kdv_oran": 0,
                    "kdv_tutar": 0,
                })

            fis_baslik = {
                "tarih": tarih,
                "fis_turu": fis_turu,
                "fis_no": satir.get("Fiş No", ""),
                "aciklama": satir.get("Açıklama", ""),
                "cari_id": cari_id,
                "toplam_tutar": tutar,
                "firma_id": firma_id,
                "yil": int(tarih[:4]),
            }

            hazir_fisler.append({
                "fis_baslik": fis_baslik,
                "fis_satirlari": fis_satirlari,
                "cek_senet_data": cek_senet_data,
                "cari_id": cari_id,
                "cari_adi": cari_adi,
                "satir_nos": [row_no],
                "fis_turu": fis_turu,
                "tarih": tarih,
                "fis_no": fis_baslik["fis_no"],
                "toplam_tutar": tutar,
                "seri_no": seri_no,
            })

    finally:
        conn.close()

    for farkli_yil in sorted(farkli_yillar):
        uyarilar.append(
            f"Seçili yıl ({aktif_yil}) dışında {farkli_yil} yılına ait satırlar bulundu. "
            f"Bu satırlar {farkli_yil} yılı olarak kaydedilecek."
        )

    return hazir_fisler, hatalar, uyarilar


def cek_senet_fislerini_kaydet(cursor, hazir_fisler, firma_id, aktif_yil):
    """Kaydeder ve yeni çek/senet ID'lerini döndürür."""
    eklenen_ids = []
    for fis in hazir_fisler:
        # 1. Çek/Senet kartını oluştur
        cek = fis["cek_senet_data"]
        cursor.execute(
            """INSERT INTO cekler_senetler (seri_no, turu, banka_id, vade_tarihi, tutar, kesideci, ciranta, aciklama, firma_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (cek["seri_no"], cek["turu"], cek["banka_id"], cek["vade"], cek["tutar"],
             cek["kesideci"], cek["ciranta"], cek["aciklama"], firma_id),
        )
        cek_id = cursor.lastrowid

        # 2. Fis satırlarındaki placeholder ID'yi güncelle
        for fs in fis["fis_satirlari"]:
            if fs["hesap_turu"] == "CekSenet" and fs["hesap_id"] == 0:
                fs["hesap_id"] = cek_id

        # 3. Fişi kaydet
        fis_id = fis_kaydet(cursor, fis["fis_baslik"], fis["fis_satirlari"], kaynak_modul="CekSenet")

        # 4. Hareket ekle
        cek_senet_hareket_ekle(
            cursor,
            cek_senet_id=cek_id,
            fis_id=fis_id,
            islem_tarihi=fis["fis_baslik"]["tarih"],
            durum=cek["durum"],
            karsi_hesap_tipi="Cari" if fis["cari_id"] else None,
            karsi_hesap_id=fis["cari_id"],
            karsi_hesap_ismi=fis.get("cari_adi"),
            aciklama=fis["fis_satirlari"][0].get("aciklama", ""),
            firma_id=firma_id,
        )

        eklenen_ids.append(fis_id)
    return eklenen_ids