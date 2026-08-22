# -*- coding: utf-8 -*-
"""
Kasa modülü Excel import yardımcıları.

- Örnek Excel şablonu oluşturur
- Excel dosyasını okur
- Satırları doğrular ve fiş/satır yapısına çevirir
- core.services.fis_kaydet ile kaydedilmeye hazır veri döndürür
"""
from datetime import datetime, date

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    load_workbook = None
    Alignment = Font = PatternFill = DataValidation = get_column_letter = None

from core.db import veritabani_baglan
from core.services import fis_kaydet


KASA_FIS_TURLERI = [
    "Kasa Gider Fişi",
    "Kasa Gelir Fişi",
    "Kasalar Arası Virman",
    "Kasa Açılış Fişi",
]

KASA_IMPORT_KOLONLARI = [
    "Fiş Türü",
    "Tarih",
    "Fiş No",
    "Genel Açıklama",
    "Kasa / Ana Kasa",
    "Hedef Kasa",
    "Gider/Gelir Kartı",
    "Yön",
    "Satır Açıklaması",
    "Miktar",
    "Birim Fiyat",
    "KDV %",
    "Tutar",
]

ZORUNLU_KOLONLAR = ["Fiş Türü", "Tarih", "Kasa / Ana Kasa"]


def _metin(deger):
    """Excel hücresini temiz metne çevirir. NaN/None ise boş string döner."""
    if deger is None:
        return ""
    if isinstance(deger, float) and pd is not None and pd.isna(deger):
        return ""
    if isinstance(deger, str):
        return deger.strip()
    if isinstance(deger, (datetime, date)):
        return deger.strftime("%d.%m.%Y")
    return str(deger).strip()


def _sayi(deger):
    """Excel hücresini float'a çevirmeyi dener. Geçersizse None döner."""
    if deger is None:
        return None
    if isinstance(deger, bool):
        return None
    if isinstance(deger, (int, float)):
        if pd is not None and isinstance(deger, float) and pd.isna(deger):
            return None
        return float(deger)

    s = str(deger).strip().replace(" ", "").replace("TL", "").replace("₺", "").replace("%", "")
    if not s:
        return None

    # "1.234,56" -> "1234.56"
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    # "1234,56" -> "1234.56"
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    # Nokta kullanımı: "1.234" binlik ayracı mı, "1234.56" ondalık mı?
    # Excel sayısal hücrelerde bu sorun olmaz; metin girişlerinde en yaygın durumları destekleyelim.
    if s.count(".") > 1:
        s = s.replace(".", "")
    elif s.count(".") == 1 and len(s.split(".")[1]) == 3:
        # Türkçe binlik ayracı varsayımı: "1.000" -> 1000, "1.234" -> 1234
        s = s.replace(".", "")

    try:
        return float(s)
    except ValueError:
        return None


def _tarih(deger):
    """Tarihi YYYY-MM-DD string'ine çevirir. Geçersizse None döner."""
    if deger is None:
        return None
    if isinstance(deger, datetime):
        return deger.strftime("%Y-%m-%d")
    if isinstance(deger, date):
        return deger.strftime("%Y-%m-%d")

    s = str(deger).strip()
    if not s:
        return None

    formatlar = [
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ]
    for fmt in formatlar:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def kasa_ornek_excel_olustur(dosya_yolu):
    """Kasa importu için örnek Excel şablonu oluşturur."""
    if pd is None:
        raise RuntimeError("Excel şablonu oluşturmak için 'pandas' ve 'openpyxl' gereklidir.")

    ornek_satirlar = [
        [
            "Kasa Gider Fişi", "01.01.2026", "G-001", "Kira ödemesi", "Ana Kasa",
            "", "Kira Gideri", "", "Ocak kirası", 1, 10000, 20, "",
        ],
        [
            "Kasa Gelir Fişi", "02.01.2026", "GL-001", "Nakit satış", "Ana Kasa",
            "", "Satış Geliri", "", "Günlük satış", 1, 5000, 20, "",
        ],
        [
            "Kasalar Arası Virman", "03.01.2026", "V-001", "Merkezden şubeye", "Ana Kasa",
            "Şube Kasası", "", "", "", "", "", "", 2000,
        ],
        [
            "Kasa Açılış Fişi", "01.01.2026", "A-001", "Açılış bakiyesi", "Ana Kasa",
            "", "", "Borç", "Açılış", "", "", "", 10000,
        ],
    ]

    df = pd.DataFrame(ornek_satirlar, columns=KASA_IMPORT_KOLONLARI)

    with pd.ExcelWriter(dosya_yolu, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Kasa İşlemleri", index=False)

    if load_workbook is None:
        return

    wb = load_workbook(dosya_yolu)
    ws = wb["Kasa İşlemleri"]

    # Fiş Türü açılır liste
    dv = DataValidation(
        type="list",
        formula1='"Kasa Gider Fişi,Kasa Gelir Fişi,Kasalar Arası Virman,Kasa Açılış Fişi"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add("A2:A2000")

    # Başlık stilleri
    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_fill = PatternFill(start_color="4A69BD", end_color="4A69BD", fill_type="solid")
    for hucre in ws[1]:
        hucre.font = baslik_font
        hucre.fill = baslik_fill
        hucre.alignment = Alignment(horizontal="center", vertical="center")

    # Sütun genişlikleri
    genislikler = {
        "A": 22, "B": 14, "C": 12, "D": 28, "E": 22,
        "F": 18, "G": 22, "H": 12, "I": 30, "J": 12,
        "K": 14, "L": 10, "M": 14,
    }
    for kolon, genislik in genislikler.items():
        ws.column_dimensions[kolon].width = genislik

    ws.freeze_panes = "A2"

    # Açıklama sayfası
    aciklama_ws = wb.create_sheet("Açıklama", 0)
    aciklama_ws.column_dimensions["A"].width = 120
    satirlar = [
        ("KASA EXCEL İMPORT REHBERİ", True),
        ("", False),
        ("1. 'Kasa İşlemleri' sayfasındaki örnek satırları kendi verilerinizle değiştirin.", False),
        ("2. 'Kasa İşlemleri' sayfasına verilerinizi girin.", False),
        ("3. Her satır bir fiş satırıdır. Aynı Fiş No'ya sahip satırlar tek fiş olarak gruplanır.", False),
        ("4. Fiş No boş bırakılırsa her Excel satırı ayrı bir fiş olarak içe aktarılır.", False),
        ("5. Kasa ve Hizmet kartı adları, Tanımlar bölümündeki adlarla birebir aynı olmalıdır.", False),
        ("6. Tanımlı olmayan kasa/kart adı girilirse import durur ve hata listesi gösterilir.", False),
        ("7. KDV % boş bırakılırsa 0 (sıfır) kabul edilir.", False),
        ("8. Gider/Gelir fişlerinde Miktar + Birim Fiyat girilebilir; sadece Tutar da girilebilir. Tutar KDV hariç ara toplamdır.", False),
        ("9. Virman fişlerinde 'Kasa / Ana Kasa' kaynak kasa, 'Hedef Kasa' hedef kasadır.", False),
        ("10. Açılış fişlerinde 'Yön' sütunu Borç veya Alacak olmalıdır.", False),
        ("", False),
        ("DESTEKLENEN FİŞ TÜRLERİ:", True),
        ("- Kasa Gider Fişi", False),
        ("- Kasa Gelir Fişi", False),
        ("- Kasalar Arası Virman", False),
        ("- Kasa Açılış Fişi", False),
        ("", False),
        ("TARİH FORMATLARI:", True),
        ("- GG.AA.YYYY  (örn: 01.01.2026)", False),
        ("- GG/AA/YYYY  (örn: 01/01/2026)", False),
        ("- YYYY-MM-DD  (örn: 2026-01-01)", False),
    ]
    for i, (metin, kalin) in enumerate(satirlar, start=1):
        hucre = aciklama_ws.cell(row=i, column=1, value=metin)
        if kalin:
            hucre.font = Font(bold=True, size=12)

    wb.save(dosya_yolu)


def kasa_excel_oku(dosya_yolu):
    """Excel dosyasındaki Kasa İşlemleri sayfasını okur ve dict listesi döndürür."""
    if pd is None:
        raise RuntimeError("Excel okumak için 'pandas' ve 'openpyxl' gereklidir.")

    try:
        df = pd.read_excel(dosya_yolu, sheet_name="Kasa İşlemleri")
    except Exception:
        # Sayfa adı farklıysa zorunlu sütunları içeren ilk sayfayı bul
        sayfalar = pd.read_excel(dosya_yolu, sheet_name=None)
        df = None
        for _, sayfa in sayfalar.items():
            if all(kolon in sayfa.columns for kolon in ZORUNLU_KOLONLAR):
                df = sayfa
                break
        if df is None:
            raise ValueError("Excel dosyasında 'Kasa İşlemleri' sayfası veya zorunlu sütunlar bulunamadı.")

    eksik = [kolon for kolon in ZORUNLU_KOLONLAR if kolon not in df.columns]
    if eksik:
        raise ValueError("Excel dosyasında şu zorunlu sütunlar bulunamadı: " + ", ".join(eksik))

    # Tamamen boş satırları at
    df = df.dropna(how="all")

    satirlar = []
    for _, row in df.iterrows():
        satir = {}
        for kolon in KASA_IMPORT_KOLONLARI:
            if kolon in df.columns:
                satir[kolon] = row.get(kolon)
            else:
                satir[kolon] = ""
        satirlar.append(satir)

    return satirlar


def _kasa_id_bul(kasa_map, kasa_adi):
    """Kasa adına göre ID döndürür. Yoksa None, çoklu kayıt varsa 'ambiguous' döner."""
    if not kasa_adi:
        return None
    kayitlar = kasa_map.get(kasa_adi, [])
    if not kayitlar:
        return None
    if len(kayitlar) > 1:
        return "ambiguous"
    return kayitlar[0]


def _hizmet_karti_bul(hizmet_map, kart_adi, tur):
    """Hizmet kartı adına göre ID döndürür. Hata durumunda string mesaj döner."""
    if not kart_adi:
        return None, "Hizmet kartı adı boş."
    kayitlar = hizmet_map.get(kart_adi, [])
    if not kayitlar:
        return None, f"'{kart_adi}' adında tanımlı bir hizmet kartı bulunamadı."
    if len(kayitlar) > 1:
        return None, f"'{kart_adi}' adında birden fazla hizmet kartı var. Lütfen kart adını netleştirin."
    kart_id, kart_turu = kayitlar[0]
    if kart_turu != tur:
        beklenen = "Gider" if tur == "Gider" else "Gelir"
        return None, f"'{kart_adi}' kartı {kart_turu} türünde; {beklenen} fişinde kullanılamaz."
    return kart_id, None


def kasa_import_dogrula(satirlar, firma_id, aktif_yil):
    """
    Excel satırlarını doğrular ve kaydedilmeye hazır fiş listesine çevirir.

    Dönüş: (hazir_fisler, hatalar, uyarilar)
    """
    hatalar = []
    uyarilar = []
    gruplar = {}

    conn = veritabani_baglan()
    try:
        cursor = conn.cursor()

        # Kasa ve hizmet kartlarını bir kez yükle
        kasa_map = {}
        cursor.execute("SELECT id, kasa_adi FROM kasalar WHERE durum=1 AND firma_id=?", (firma_id,))
        for kart_id, kasa_adi in cursor.fetchall():
            kasa_map.setdefault(kasa_adi, []).append(kart_id)

        hizmet_map = {}
        cursor.execute("SELECT id, kart_adi, tur FROM hizmet_kartlari WHERE durum=1 AND firma_id=?", (firma_id,))
        for kart_id, kart_adi, kart_turu in cursor.fetchall():
            hizmet_map.setdefault(kart_adi, []).append((kart_id, kart_turu))

        for index, raw_satir in enumerate(satirlar, start=2):
            row_no = index  # Excel'de 1. satır başlık olduğu için 2'den başlatıyoruz

            # Hücreleri normalize et
            satir = {}
            for kolon in KASA_IMPORT_KOLONLARI:
                deger = raw_satir.get(kolon, "") if isinstance(raw_satir, dict) else raw_satir
                if kolon in ("Miktar", "Birim Fiyat", "KDV %", "Tutar"):
                    satir[kolon] = _sayi(deger)
                elif kolon == "Tarih":
                    satir[kolon] = _tarih(deger)
                else:
                    satir[kolon] = _metin(deger)

            fis_turu = satir.get("Fiş Türü", "")
            if not fis_turu:
                hatalar.append(f"Satır {row_no}: Fiş Türü boş bırakılamaz.")
                continue
            if fis_turu not in KASA_FIS_TURLERI:
                hatalar.append(f"Satır {row_no}: Geçersiz Fiş Türü '{fis_turu}'.")
                continue

            tarih = satir.get("Tarih")
            if not tarih:
                hatalar.append(f"Satır {row_no}: Tarih geçersiz veya boş.")
                continue

            kasa_adi = satir.get("Kasa / Ana Kasa", "")
            if not kasa_adi:
                hatalar.append(f"Satır {row_no}: 'Kasa / Ana Kasa' boş bırakılamaz.")
                continue

            kasa_id = _kasa_id_bul(kasa_map, kasa_adi)
            if kasa_id is None:
                hatalar.append(f"Satır {row_no}: '{kasa_adi}' adında tanımlı kasa bulunamadı.")
                continue
            if kasa_id == "ambiguous":
                hatalar.append(f"Satır {row_no}: '{kasa_adi}' adında birden fazla kasa var. Lütfen kasa adını netleştirin.")
                continue

            fis_no = satir.get("Fiş No", "")
            genel_aciklama = satir.get("Genel Açıklama", "")
            satir_aciklama = satir.get("Satır Açıklaması", "")

            # Açılış fişinde kasa satır bazındadır; anahtar fiş no üzerinden kurulur.
            if fis_turu == "Kasa Açılış Fişi":
                yon = satir.get("Yön", "")
                if yon not in ("Borç", "Alacak"):
                    hatalar.append(f"Satır {row_no}: Yön 'Borç' veya 'Alacak' olmalıdır.")
                    continue

                tutar = satir.get("Tutar")
                if tutar is None or tutar <= 0:
                    hatalar.append(f"Satır {row_no}: Açılış tutarı 0'dan büyük olmalıdır.")
                    continue

                line = {
                    "hesap_turu": "Kasa",
                    "hesap_id": kasa_id,
                    "borc": tutar if yon == "Borç" else 0,
                    "alacak": tutar if yon == "Alacak" else 0,
                    "aciklama": satir_aciklama,
                    "miktar": 1,
                    "birim_fiyat": tutar,
                    "kdv_oran": 0,
                    "kdv_tutar": 0,
                }
                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", None)
                _gruba_ekle(gruplar, anahtar, line, row_no, fis_turu, tarih, fis_no, genel_aciklama, kasa_adi, None, tutar)

            elif fis_turu == "Kasalar Arası Virman":
                hedef_kasa_adi = satir.get("Hedef Kasa", "")
                if not hedef_kasa_adi:
                    hatalar.append(f"Satır {row_no}: Virman için Hedef Kasa boş bırakılamaz.")
                    continue

                hedef_kasa_id = _kasa_id_bul(kasa_map, hedef_kasa_adi)
                if hedef_kasa_id is None:
                    hatalar.append(f"Satır {row_no}: '{hedef_kasa_adi}' adında tanımlı kasa bulunamadı.")
                    continue
                if hedef_kasa_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{hedef_kasa_adi}' adında birden fazla kasa var. Lütfen kasa adını netleştirin.")
                    continue
                if kasa_id == hedef_kasa_id:
                    hatalar.append(f"Satır {row_no}: Virman işleminde kaynak ve hedef kasa aynı olamaz.")
                    continue

                tutar = satir.get("Tutar")
                if tutar is None or tutar <= 0:
                    hatalar.append(f"Satır {row_no}: Virman tutarı 0'dan büyük olmalıdır.")
                    continue

                # Virman tek satırlık bir fiştir; aynı Fiş No ile tekrar edemez.
                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", None)
                if anahtar in gruplar:
                    hatalar.append(f"Satır {row_no}: Virman fişinde aynı Fiş No ile birden fazla satır bulunamaz.")
                    continue

                line1 = {
                    "hesap_turu": "Kasa",
                    "hesap_id": kasa_id,
                    "borc": 0,
                    "alacak": tutar,
                    "aciklama": f"ID:{hedef_kasa_id} kasaya virman",
                    "miktar": 1,
                    "birim_fiyat": tutar,
                    "kdv_oran": 0,
                    "kdv_tutar": 0,
                }
                line2 = {
                    "hesap_turu": "Kasa",
                    "hesap_id": hedef_kasa_id,
                    "borc": tutar,
                    "alacak": 0,
                    "aciklama": f"ID:{kasa_id} kasadan virman",
                    "miktar": 1,
                    "birim_fiyat": tutar,
                    "kdv_oran": 0,
                    "kdv_tutar": 0,
                }
                _gruba_ekle(gruplar, anahtar, line1, row_no, fis_turu, tarih, fis_no, genel_aciklama, kasa_adi, hedef_kasa_adi, tutar)
                _gruba_ekle(gruplar, anahtar, line2, row_no, fis_turu, tarih, fis_no, genel_aciklama, kasa_adi, hedef_kasa_adi, tutar)
                # Virman iki satırdan oluşur; toplam tutar tek satır değeri olmalıdır.
                gruplar[anahtar]["toplam_tutar"] = tutar
                gruplar[anahtar]["ana_kasa_id"] = kasa_id

            else:  # Kasa Gider / Gelir
                is_gider = fis_turu == "Kasa Gider Fişi"
                beklenen_tur = "Gider" if is_gider else "Gelir"
                hesap_adi = satir.get("Gider/Gelir Kartı", "")
                if not hesap_adi:
                    hatalar.append(f"Satır {row_no}: Gider/Gelir kartı boş bırakılamaz.")
                    continue

                hesap_id, hata = _hizmet_karti_bul(hizmet_map, hesap_adi, beklenen_tur)
                if hata:
                    hatalar.append(f"Satır {row_no}: {hata}")
                    continue

                miktar = satir.get("Miktar")
                birim_fiyat = satir.get("Birim Fiyat")
                tutar = satir.get("Tutar")
                kdv_oran = satir.get("KDV %")
                if kdv_oran is None:
                    kdv_oran = 0.0

                # Esnek tutar girişi:
                # - Miktar + Birim Fiyat varsa onlar kullanılır.
                # - Sadece Tutar varsa miktar 1, birim fiyat tutar kabul edilir.
                if miktar is None:
                    miktar = 1
                if birim_fiyat is None and tutar is not None:
                    if miktar and miktar > 0:
                        birim_fiyat = tutar / miktar
                    else:
                        hatalar.append(f"Satır {row_no}: Miktar 0 olduğu için birim fiyat hesaplanamadı.")
                        continue
                if birim_fiyat is None:
                    hatalar.append(f"Satır {row_no}: Birim Fiyat veya Tutar girilmelidir.")
                    continue

                if miktar <= 0 or birim_fiyat <= 0:
                    hatalar.append(f"Satır {row_no}: Miktar ve Birim Fiyat 0'dan büyük olmalıdır.")
                    continue

                if tutar is not None and abs(miktar * birim_fiyat - tutar) > 0.01:
                    uyarilar.append(
                        f"Satır {row_no}: Girilen Tutar ile Miktar*Birim Fiyat farklı. "
                        "Miktar*Birim Fiyat esas alındı."
                    )

                ara_toplam = miktar * birim_fiyat
                kdv_tutar = ara_toplam * (kdv_oran / 100.0)
                genel_toplam = ara_toplam + kdv_tutar

                line = {
                    "hesap_turu": "Hizmet",
                    "hesap_id": hesap_id,
                    "borc": ara_toplam if is_gider else 0,
                    "alacak": 0 if is_gider else ara_toplam,
                    "aciklama": satir_aciklama,
                    "miktar": miktar,
                    "birim_fiyat": birim_fiyat,
                    "kdv_oran": kdv_oran,
                    "kdv_tutar": kdv_tutar,
                }
                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", kasa_id)
                _gruba_ekle(gruplar, anahtar, line, row_no, fis_turu, tarih, fis_no, genel_aciklama, kasa_adi, None, genel_toplam)
                gruplar[anahtar]["ana_kasa_id"] = kasa_id

    finally:
        conn.close()

    hazir_fisler = []
    for grup in gruplar.values():
        if not grup["fis_satirlari"]:
            continue

        fis_turu = grup["fis_turu"]
        fis_satirlari = grup["fis_satirlari"]

        if fis_turu in ("Kasa Gider Fişi", "Kasa Gelir Fişi"):
            is_gider = fis_turu == "Kasa Gider Fişi"
            toplam = grup["toplam_tutar"]
            kasa_karsi = {
                "hesap_turu": "Kasa",
                "hesap_id": grup["ana_kasa_id"],
                "borc": 0 if is_gider else toplam,
                "alacak": toplam if is_gider else 0,
                "aciklama": grup["genel_aciklama"],
                "miktar": 1,
                "birim_fiyat": toplam,
                "kdv_oran": 0,
                "kdv_tutar": 0,
            }
            fis_satirlari = fis_satirlari + [kasa_karsi]

        fis_baslik = {
            "tarih": grup["tarih"],
            "fis_turu": fis_turu,
            "fis_no": grup["fis_no"],
            "aciklama": grup["genel_aciklama"],
            "toplam_tutar": grup["toplam_tutar"],
            "cari_id": None,
            "firma_id": firma_id,
            "yil": aktif_yil,
        }

        hazir_fisler.append({
            "fis_baslik": fis_baslik,
            "fis_satirlari": fis_satirlari,
            "satir_nos": grup["satir_nos"],
            "fis_turu": fis_turu,
            "tarih": grup["tarih"],
            "fis_no": grup["fis_no"],
            "kasa_adi": grup["kasa_adi"],
            "hedef_kasa_adi": grup.get("hedef_kasa_adi", ""),
            "toplam_tutar": grup["toplam_tutar"],
            "satir_sayisi": len(grup["fis_satirlari"]),
        })

    return hazir_fisler, hatalar, uyarilar


def _gruba_ekle(gruplar, anahtar, line, row_no, fis_turu, tarih, fis_no, genel_aciklama, kasa_adi, hedef_kasa_adi, tutar):
    """Aynı fiş anahtarına satır ekler; grup yoksa oluşturur."""
    if anahtar not in gruplar:
        gruplar[anahtar] = {
            "fis_turu": fis_turu,
            "tarih": tarih,
            "fis_no": fis_no or "",
            "genel_aciklama": genel_aciklama,
            "ana_kasa_id": None,
            "kasa_adi": kasa_adi,
            "hedef_kasa_adi": hedef_kasa_adi or "",
            "fis_satirlari": [],
            "satir_nos": [],
            "toplam_tutar": 0.0,
        }

    grup = gruplar[anahtar]

    # Genel açıklama ilk dolu değerle doldurulur
    if not grup["genel_aciklama"] and genel_aciklama:
        grup["genel_aciklama"] = genel_aciklama

    grup["fis_satirlari"].append(line)
    if row_no not in grup["satir_nos"]:
        grup["satir_nos"].append(row_no)
    grup["toplam_tutar"] = grup.get("toplam_tutar", 0.0) + tutar


def kasa_fislerini_kaydet(cursor, hazir_fisler, firma_id, aktif_yil):
    """Doğrulanmış fiş listesini veritabanına kaydeder ve eklenen fiş ID'lerini döndürür."""
    eklenen_ids = []
    for fis in hazir_fisler:
        yeni_id = fis_kaydet(cursor, fis["fis_baslik"], fis["fis_satirlari"], kaynak_modul="Kasa")
        eklenen_ids.append(yeni_id)
    return eklenen_ids
