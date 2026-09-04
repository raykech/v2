# -*- coding: utf-8 -*-
"""
Cari modülü Excel import yardımcıları.
"""
from datetime import datetime, date

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    load_workbook = None
    Alignment = Font = PatternFill = DataValidation = None

from core.db import veritabani_baglan
from core.services import fis_kaydet


CARI_FIS_TURLERI = [
    "Alacak Dekontu",
    "Borç Dekontu",
    "Cari Ödeme",
    "Cari Tahsilat",
    "Cari Virman",
]

CARI_IMPORT_KOLONLARI = [
    "Fiş Türü",
    "Tarih",
    "Fiş No",
    "Açıklama",
    "Ana Kart / Ödeme Hesabı",
    "Ödeme Türü",
    "Cari",
    "Yön",
    "Satır Açıklaması",
    "Tutar",
]

ZORUNLU_KOLONLAR = ["Fiş Türü", "Tarih"]





from utils.import_helpers import metin as _metin, sayi as _sayi, tarih as _tarih, cari_id_bul as _cari_id_bul, hizmet_id_bul as _hizmet_id_bul, kasa_id_bul as _kasa_id_bul, banka_id_bul as _banka_id_bul

def cari_ornek_excel_olustur(dosya_yolu):
    if pd is None:
        raise RuntimeError("Excel şablonu oluşturmak için 'pandas' ve 'openpyxl' gereklidir.")
    ornek = [
        # Alacak Dekontu
        ["Alacak Dekontu", "01.01.2026", "AD-001", "Alacak dekontu", "Kira Gideri", "", "Müşteri A", "", "Ocak", 5000],
        # Borç Dekontu
        ["Borç Dekontu", "02.01.2026", "BD-001", "Borç dekontu", "İskonto Geliri", "", "Tedarikçi B", "", "", 2000],
        # Cari Ödeme
        ["Cari Ödeme", "03.01.2026", "CO-001", "Ödeme", "TL KASASI", "Kasa", "Tedarikçi B", "", "Fatura", 3000],
        # Cari Tahsilat
        ["Cari Tahsilat", "04.01.2026", "CT-001", "Tahsilat", "TL KASASI", "Banka", "Müşteri A", "", "", 1500],
        # Cari Virman
        ["Cari Virman", "05.01.2026", "CV-001", "Virman", "", "", "Müşteri A", "Borç", "Virman", 1000],
        ["Cari Virman", "05.01.2026", "CV-001", "Virman", "", "", "Tedarikçi B", "Alacak", "Virman", 1000],
    ]
    df = pd.DataFrame(ornek, columns=CARI_IMPORT_KOLONLARI)
    with pd.ExcelWriter(dosya_yolu, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Cari İşlemleri", index=False)
    if load_workbook is None:
        return
    wb = load_workbook(dosya_yolu)
    ws = wb["Cari İşlemleri"]
    dv = DataValidation(type="list", formula1='"Alacak Dekontu,Borç Dekontu,Cari Ödeme,Cari Tahsilat,Cari Virman"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add("A2:A2000")
    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_fill = PatternFill(start_color="4A69BD", end_color="4A69BD", fill_type="solid")
    for hucre in ws[1]:
        hucre.font = baslik_font; hucre.fill = baslik_fill; hucre.alignment = Alignment(horizontal="center", vertical="center")
    for c, w in [("A", 22), ("B", 14), ("C", 12), ("D", 28), ("E", 24), ("F", 14), ("G", 22), ("H", 12), ("I", 30), ("J", 14)]:
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A2"
    aciklama_ws = wb.create_sheet("Açıklama", 0)
    aciklama_ws.column_dimensions["A"].width = 120
    satirlar = [
        ("CARİ EXCEL İMPORT REHBERİ", True),
        ("", False),
        ("Alacak Dekontu: Ana Kart / Ödeme Hesabı'na Gider kartı girin, satırlara Cariler.", False),
        ("Borç Dekontu: Ana Kart / Ödeme Hesabı'na Gelir kartı girin, satırlara Cariler.", False),
        ("Cari Ödeme/Tahsilat: Ödeme Türü (Kasa/Banka) ve Ödeme Hesabı girin, satırlara Cariler.", False),
        ("Örnekteki TL KASA ve Banka Hesabı 1 adlarını kendi kasa/banka tanımlarınızla değiştirin.", False),
        ("Cari Virman: Satırlara Yön (Borç/Alacak) ile Cariler ve Tutar girin, toplam eşit olmalı.", False),
        ("", False),
        ("DESTEKLENEN FİŞ TÜRLERİ:", True),
        ("- Alacak Dekontu", False),
        ("- Borç Dekontu", False),
        ("- Cari Ödeme", False),
        ("- Cari Tahsilat", False),
        ("- Cari Virman", False),
    ]
    for i, (metin, kalin) in enumerate(satirlar, start=1):
        hucre = aciklama_ws.cell(row=i, column=1, value=metin)
        if kalin: hucre.font = Font(bold=True, size=12)
    wb.save(dosya_yolu)


def cari_excel_oku(dosya_yolu):
    if pd is None:
        raise RuntimeError("Excel okumak için 'pandas' ve 'openpyxl' gereklidir.")
    try:
        df = pd.read_excel(dosya_yolu, sheet_name="Cari İşlemleri")
    except Exception:
        sayfalar = pd.read_excel(dosya_yolu, sheet_name=None)
        df = None
        for _, sayfa in sayfalar.items():
            if all(kolon in sayfa.columns for kolon in ZORUNLU_KOLONLAR):
                df = sayfa; break
        if df is None:
            raise ValueError("Excel dosyasında 'Cari İşlemleri' sayfası veya zorunlu sütunlar bulunamadı.")
    eksik = [kolon for kolon in ZORUNLU_KOLONLAR if kolon not in df.columns]
    if eksik:
        raise ValueError("Excel dosyasında şu zorunlu sütunlar bulunamadı: " + ", ".join(eksik))
    df = df.dropna(how="all")
    satirlar = []
    for _, row in df.iterrows():
        satir = {}
        for kolon in CARI_IMPORT_KOLONLARI:
            satir[kolon] = row.get(kolon) if kolon in df.columns else ""
        satirlar.append(satir)
    return satirlar






def _satir_dik(hesap_turu, hesap_id, borc, alacak, aciklama, tutar):
    return {
        "hesap_turu": hesap_turu,
        "hesap_id": hesap_id,
        "borc": borc,
        "alacak": alacak,
        "aciklama": aciklama,
        "miktar": 1,
        "birim_fiyat": tutar,
        "kdv_oran": 0,
        "kdv_tutar": 0,
    }


def cari_import_dogrula(satirlar, firma_id, aktif_yil):
    hatalar = []
    uyarilar = []
    farkli_yillar = set()
    gruplar = {}

    conn = veritabani_baglan()
    try:
        cursor = conn.cursor()

        for index, raw_satir in enumerate(satirlar, start=2):
            row_no = index
            satir = {}
            for kolon in CARI_IMPORT_KOLONLARI:
                deger = raw_satir.get(kolon, "") if isinstance(raw_satir, dict) else raw_satir
                if kolon == "Tutar":
                    satir[kolon] = _sayi(deger)
                elif kolon == "Tarih":
                    satir[kolon] = _tarih(deger)
                else:
                    satir[kolon] = _metin(deger)

            fis_turu = satir.get("Fiş Türü", "")
            if not fis_turu:
                hatalar.append(f"Satır {row_no}: Fiş Türü boş bırakılamaz.")
                continue
            if fis_turu not in CARI_FIS_TURLERI:
                hatalar.append(f"Satır {row_no}: Geçersiz Fiş Türü '{fis_turu}'.")
                continue

            tarih = satir.get("Tarih")
            if not tarih:
                hatalar.append(f"Satır {row_no}: Tarih geçersiz veya boş.")
                continue

            if int(tarih[:4]) != aktif_yil:
                farkli_yillar.add(int(tarih[:4]))

            fis_no = satir.get("Fiş No", "")
            genel_aciklama = satir.get("Açıklama", "")
            satir_aciklama = satir.get("Satır Açıklaması", "")
            cari_adi = satir.get("Cari", "")
            yon = satir.get("Yön", "")

            if fis_turu in ("Alacak Dekontu", "Borç Dekontu"):
                ana_kart_adi = satir.get("Ana Kart / Ödeme Hesabı", "")
                if not ana_kart_adi:
                    hatalar.append(f"Satır {row_no}: Alacak/Borç Dekontu için Ana Kart gerekli.")
                    continue
                beklenen_tur = "Gider" if fis_turu == "Alacak Dekontu" else "Gelir"
                ana_kart_id = _hizmet_id_bul(cursor, ana_kart_adi, beklenen_tur, firma_id)
                if ana_kart_id is None:
                    hatalar.append(f"Satır {row_no}: '{ana_kart_adi}' adında uygun hizmet kartı bulunamadı.")
                    continue
                if ana_kart_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{ana_kart_adi}' adında birden fazla hizmet kartı var.")
                    continue
                if ana_kart_id == "wrong_type":
                    hatalar.append(f"Satır {row_no}: '{ana_kart_adi}' kartı bu fiş türü için uygun değil.")
                    continue

                if not cari_adi:
                    hatalar.append(f"Satır {row_no}: Cari boş bırakılamaz.")
                    continue
                cari_id = _cari_id_bul(cursor, cari_adi, firma_id)
                if cari_id is None:
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında cari bulunamadı.")
                    continue
                if cari_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında birden fazla cari var.")
                    continue

                tutar = satir.get("Tutar")
                if tutar is None or tutar <= 0:
                    hatalar.append(f"Satır {row_no}: Tutar 0'dan büyük olmalıdır.")
                    continue

                # Alacak Dekontu: cariler alacaklı, Borç Dekontu: cariler borçlu
                cari_borclu = fis_turu == "Borç Dekontu"
                line = _satir_dik("Cari", cari_id, tutar if cari_borclu else 0, 0 if cari_borclu else tutar, satir_aciklama, tutar)
                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", ana_kart_id, "", "")
                if anahtar not in gruplar:
                    gruplar[anahtar] = {
                        "fis_turu": fis_turu, "tarih": tarih, "fis_no": fis_no or "",
                        "genel_aciklama": genel_aciklama, "ana_kart_id": ana_kart_id,
                        "odeme_turu": "", "odeme_hesap_id": None,
                        "fis_satirlari": [], "satir_nos": [], "toplam_tutar": 0.0,
                    }
                grup = gruplar[anahtar]
                if not grup["genel_aciklama"] and genel_aciklama:
                    grup["genel_aciklama"] = genel_aciklama
                grup["fis_satirlari"].append(line)
                if row_no not in grup["satir_nos"]:
                    grup["satir_nos"].append(row_no)
                grup["toplam_tutar"] += tutar

            elif fis_turu in ("Cari Ödeme", "Cari Tahsilat"):
                odeme_turu = satir.get("Ödeme Türü", "")
                if odeme_turu not in ("Kasa", "Banka"):
                    hatalar.append(f"Satır {row_no}: Ödeme Türü Kasa veya Banka olmalıdır.")
                    continue
                odeme_hesap_adi = satir.get("Ana Kart / Ödeme Hesabı", "")
                if not odeme_hesap_adi:
                    hatalar.append(f"Satır {row_no}: Ödeme Hesabı boş bırakılamaz.")
                    continue
                if odeme_turu == "Kasa":
                    odeme_hesap_id = _kasa_id_bul(cursor, odeme_hesap_adi, firma_id)
                else:
                    odeme_hesap_id = _banka_id_bul(cursor, odeme_hesap_adi, firma_id)
                if odeme_hesap_id is None:
                    hatalar.append(f"Satır {row_no}: '{odeme_hesap_adi}' adında {odeme_turu} hesabı bulunamadı.")
                    continue
                if odeme_hesap_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{odeme_hesap_adi}' adında birden fazla hesap var.")
                    continue

                if not cari_adi:
                    hatalar.append(f"Satır {row_no}: Cari boş bırakılamaz.")
                    continue
                cari_id = _cari_id_bul(cursor, cari_adi, firma_id)
                if cari_id is None:
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında cari bulunamadı.")
                    continue
                if cari_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında birden fazla cari var.")
                    continue

                tutar = satir.get("Tutar")
                if tutar is None or tutar <= 0:
                    hatalar.append(f"Satır {row_no}: Tutar 0'dan büyük olmalıdır.")
                    continue

                # Cari Ödeme: cariler borçlu, karşı kasa/banka alacaklı
                # Cari Tahsilat: cariler alacaklı, karşı kasa/banka borçlu
                cari_borclu = fis_turu == "Cari Ödeme"
                line = _satir_dik("Cari", cari_id, tutar if cari_borclu else 0, 0 if cari_borclu else tutar, satir_aciklama, tutar)
                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", odeme_hesap_id, odeme_turu, "")
                if anahtar not in gruplar:
                    gruplar[anahtar] = {
                        "fis_turu": fis_turu, "tarih": tarih, "fis_no": fis_no or "",
                        "genel_aciklama": genel_aciklama, "ana_kart_id": None,
                        "odeme_turu": odeme_turu, "odeme_hesap_id": odeme_hesap_id,
                        "fis_satirlari": [], "satir_nos": [], "toplam_tutar": 0.0,
                    }
                grup = gruplar[anahtar]
                if not grup["genel_aciklama"] and genel_aciklama:
                    grup["genel_aciklama"] = genel_aciklama
                grup["fis_satirlari"].append(line)
                if row_no not in grup["satir_nos"]:
                    grup["satir_nos"].append(row_no)
                grup["toplam_tutar"] += tutar

            elif fis_turu == "Cari Virman":
                if not cari_adi:
                    hatalar.append(f"Satır {row_no}: Cari boş bırakılamaz.")
                    continue
                cari_id = _cari_id_bul(cursor, cari_adi, firma_id)
                if cari_id is None:
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında cari bulunamadı.")
                    continue
                if cari_id == "ambiguous":
                    hatalar.append(f"Satır {row_no}: '{cari_adi}' adında birden fazla cari var.")
                    continue
                if yon not in ("Borç", "Alacak"):
                    hatalar.append(f"Satır {row_no}: Yön 'Borç' veya 'Alacak' olmalıdır.")
                    continue
                tutar = satir.get("Tutar")
                if tutar is None or tutar <= 0:
                    hatalar.append(f"Satır {row_no}: Tutar 0'dan büyük olmalıdır.")
                    continue

                line = _satir_dik("Cari", cari_id, tutar if yon == "Borç" else 0, 0 if yon == "Borç" else tutar, satir_aciklama, tutar)
                anahtar = (fis_turu, tarih, fis_no or f"#SATIR{row_no}", None, "", "")
                if anahtar not in gruplar:
                    gruplar[anahtar] = {
                        "fis_turu": fis_turu, "tarih": tarih, "fis_no": fis_no or "",
                        "genel_aciklama": genel_aciklama, "ana_kart_id": None,
                        "odeme_turu": "", "odeme_hesap_id": None,
                        "fis_satirlari": [], "satir_nos": [], "toplam_tutar": 0.0,
                    }
                grup = gruplar[anahtar]
                if not grup["genel_aciklama"] and genel_aciklama:
                    grup["genel_aciklama"] = genel_aciklama
                grup["fis_satirlari"].append(line)
                if row_no not in grup["satir_nos"]:
                    grup["satir_nos"].append(row_no)
                grup["toplam_tutar"] += tutar

    finally:
        conn.close()

    hazir_fisler = []
    for grup in gruplar.values():
        if not grup["fis_satirlari"]:
            continue
        fis_turu = grup["fis_turu"]
        fis_satirlari = list(grup["fis_satirlari"])
        toplam = grup["toplam_tutar"]

        # Cari Virman: borç = alacak kontrolü
        if fis_turu == "Cari Virman":
            borc = sum(s['borc'] for s in fis_satirlari)
            alacak = sum(s['alacak'] for s in fis_satirlari)
            if abs(borc - alacak) > 0.005:
                hatalar.append(f"Virman fişi (No: {grup['fis_no']}): Borç toplamı ({borc:.2f}) ile alacak toplamı ({alacak:.2f}) eşit değil.")
                continue

        # Alacak/Borç Dekontu: karşı satır ekle
        if fis_turu in ("Alacak Dekontu", "Borç Dekontu"):
            ana_kart_id = grup["ana_kart_id"]
            if fis_turu == "Alacak Dekontu":
                karsi = _satir_dik("Hizmet", ana_kart_id, toplam, 0, "Alacak Dekontu", toplam)
            else:
                karsi = _satir_dik("Hizmet", ana_kart_id, 0, toplam, "Borç Dekontu", toplam)
            fis_satirlari.append(karsi)

        # Cari Ödeme / Tahsilat: karşı satır ekle
        if fis_turu in ("Cari Ödeme", "Cari Tahsilat"):
            odeme_hesap_id = grup["odeme_hesap_id"]
            odeme_hesap_turu = "Banka" if grup["odeme_turu"] == "Banka" else "Kasa"
            if fis_turu == "Cari Ödeme":
                karsi = _satir_dik(odeme_hesap_turu, odeme_hesap_id, 0, toplam, "Cari Ödeme", toplam)
            else:
                karsi = _satir_dik(odeme_hesap_turu, odeme_hesap_id, toplam, 0, "Cari Tahsilat", toplam)
            fis_satirlari.append(karsi)

        fis_baslik = {
            "tarih": grup["tarih"],
            "fis_turu": fis_turu,
            "fis_no": grup["fis_no"],
            "aciklama": grup["genel_aciklama"],
            "cari_id": None,
            "toplam_tutar": toplam,
            "firma_id": firma_id,
            "yil": int(grup["tarih"][:4]),
        }

        hazir_fisler.append({
            "fis_baslik": fis_baslik,
            "fis_satirlari": fis_satirlari,
            "satir_nos": grup["satir_nos"],
            "fis_turu": fis_turu,
            "tarih": grup["tarih"],
            "fis_no": grup["fis_no"],
            "cari_adi": "",
            "odeme_turu": grup.get("odeme_turu", ""),
            "odeme_hesap_adi": "",
            "toplam_tutar": toplam,
        })

    for farkli_yil in sorted(farkli_yillar):
        uyarilar.append(
            f"Seçili yıl ({aktif_yil}) dışında {farkli_yil} yılına ait satırlar bulundu. "
            f"Bu satırlar {farkli_yil} yılı olarak kaydedilecek."
        )

    return hazir_fisler, hatalar, uyarilar


def cari_fislerini_kaydet(cursor, hazir_fisler, firma_id, aktif_yil):
    eklenen_ids = []
    for fis in hazir_fisler:
        yeni_id = fis_kaydet(cursor, fis["fis_baslik"], fis["fis_satirlari"], kaynak_modul="Cari")
        eklenen_ids.append(yeni_id)
    return eklenen_ids